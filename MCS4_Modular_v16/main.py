import csv
import json
import math
import random
import sys
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path

import pyqtgraph as pg
from PySide6.QtCore import QTimer, Qt
from PySide6.QtGui import QTextDocument
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QPushButton,
    QPlainTextEdit,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

try:
    import serial
    from serial.tools import list_ports
except Exception:
    serial = None
    list_ports = None


APP_VERSION = "1.5"
SYNC_BYTE = 0xFF
PACKET_LENGTH = 8


REC_DIR = Path("recordings")
REC_DIR.mkdir(exist_ok=True)
EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)
CONFIG_FILE = Path("sensor_config.json")


def hex_string(data: bytes) -> str:
    return " ".join(f"{b:02X}" for b in data)


def make_packet(channel: int, unit_code: int, value: int) -> bytes:
    value = max(0, min(value, 0x3FFF))
    return bytes([
        SYNC_BYTE,
        0x00,
        0x01,
        0x02,
        channel & 0xFF,
        unit_code & 0x0F,
        (value >> 8) & 0x3F,
        value & 0xFF,
    ])


def default_sensor_config() -> dict:
    return {
        "sensors": [
            {"id": 1, "name": "Öltemperatur", "unit_code": 1, "warn_low": 0, "warn_high": 90, "alarm_low": -10, "alarm_high": 95},
            {"id": 2, "name": "Öldruck", "unit_code": 2, "warn_low": 4.5, "warn_high": 6.0, "alarm_low": 4.0, "alarm_high": 6.5},
            {"id": 3, "name": "Drehzahl", "unit_code": 3, "warn_low": 500, "warn_high": 1900, "alarm_low": 300, "alarm_high": 2100},
            {"id": 4, "name": "Kühlwasser", "unit_code": 1, "warn_low": 0, "warn_high": 88, "alarm_low": -10, "alarm_high": 95},
            {"id": 5, "name": "Abgastemperatur", "unit_code": 1, "warn_low": 0, "warn_high": 520, "alarm_low": -10, "alarm_high": 560}
        ]
    }


def ensure_sensor_config() -> dict:
    if not CONFIG_FILE.exists():
        CONFIG_FILE.write_text(json.dumps(default_sensor_config(), indent=2, ensure_ascii=False), encoding="utf-8")
    with CONFIG_FILE.open("r", encoding="utf-8") as f:
        return json.load(f)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"MCS-4 Monitor - Version {APP_VERSION} - MCS-4 Decoder Basis")
        self.resize(1360, 860)

        self.t = 0.0
        self.packet_count = 0
        self.byte_count = 0
        self.serial_port = None
        self.buffer = bytearray()

        self.recording = False
        self.record_start_time: datetime | None = None
        self.record_file = None
        self.record_file_path: Path | None = None
        self.record_packet_count = 0
        self.rec_blink = False

        self.player_packets: list[bytes] = []
        self.player_index = 0
        self.player_file_path: Path | None = None

        self.config_data = ensure_sensor_config()

        # Vorläufige Einheitentabelle. Diese wird mit den Tabellen aus der
        # MCS-4-Dokumentation schrittweise vervollständigt.
        self.units = {
            0: "",
            1: "°C",
            2: "bar",
            3: "rpm",
            4: "V",
            5: "A",
            6: "%",
        }

        self.sensor_names = {}
        self.sensor_unit_codes = {}
        self.limits = {}
        self.load_sensor_config_from_file()

        # Vorläufige Word-Type-Benennung. Die Bitbelegung wird in den nächsten
        # Schritten mit der PDF abgeglichen und erweitert.
        self.word_types = {
            0: "Data Value / Messwert",
            1: "Limit Value / Grenzwert",
            2: "Alarm Message / Alarmmeldung",
            3: "Binary Signal / Binärsignal",
            4: "Status Message / Status",
            5: "Curve Transfer / Kurve",
            6: "Control Command / Steuerbefehl",
            7: "Key Identification / Kennung",
        }

        self.sensor_values = {}
        self.active_alarms = {}
        self.trend_data = defaultdict(lambda: deque(maxlen=250))
        self.curves = {}

        self._build_ui()
        self.refresh_ports()

        self.sim_timer = QTimer(self)
        self.sim_timer.timeout.connect(self.update_simulator)

        self.serial_timer = QTimer(self)
        self.serial_timer.timeout.connect(self.read_serial)

        self.player_timer = QTimer(self)
        self.player_timer.timeout.connect(self.player_tick)

        self.diag_timer = QTimer(self)
        self.diag_timer.timeout.connect(self.update_diagnostics)
        self.diag_timer.start(1000)

        self.rec_timer = QTimer(self)
        self.rec_timer.timeout.connect(self.update_recording_status)
        self.rec_timer.start(500)

    def _build_ui(self):
        root = QWidget()
        main_layout = QVBoxLayout(root)

        title_row = QHBoxLayout()
        title = QLabel("MCS-4 Professional Monitor")
        title.setStyleSheet("font-size: 22px; font-weight: bold;")
        version = QLabel(f"Version {APP_VERSION}")
        version.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        title_row.addWidget(title)
        title_row.addStretch()
        title_row.addWidget(version)
        main_layout.addLayout(title_row)

        header = QFrame()
        header.setFrameShape(QFrame.StyledPanel)
        header.setStyleSheet("QFrame { background: #f4f6f8; border-radius: 6px; }")
        header_layout = QGridLayout(header)

        self.mode = QComboBox()
        self.mode.addItems(["Simulator", "RS422", "Player"])
        self.port_box = QComboBox()
        self.refresh_btn = QPushButton("Ports aktualisieren")
        self.start_btn = QPushButton("Start")
        self.stop_btn = QPushButton("Stop")
        self.record_start_btn = QPushButton("Aufnahme Start")
        self.record_stop_btn = QPushButton("Aufnahme Stop")
        self.player_load_btn = QPushButton("Player-Datei laden")
        self.export_snapshot_btn = QPushButton("Export CSV")
        self.export_excel_btn = QPushButton("Export Excel")
        self.export_pdf_btn = QPushButton("Export PDF-Bericht")
        self.export_trend_btn = QPushButton("Export Trend PNG")

        self.status = QLabel("Status: bereit")
        self.status.setStyleSheet("font-weight: bold;")

        self.rec_label = QLabel("⚫ REC AUS")
        self.rec_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #333;")
        self.rec_time_label = QLabel("00:00:00")
        self.rec_count_label = QLabel("Telegramme: 0")
        self.rec_file_label = QLabel("Datei: -")

        header_layout.addWidget(QLabel("Modus:"), 0, 0)
        header_layout.addWidget(self.mode, 0, 1)
        header_layout.addWidget(QLabel("COM-Port:"), 0, 2)
        header_layout.addWidget(self.port_box, 0, 3)
        header_layout.addWidget(self.refresh_btn, 0, 4)
        header_layout.addWidget(self.player_load_btn, 0, 5)
        header_layout.addWidget(self.export_snapshot_btn, 2, 0)
        header_layout.addWidget(self.export_excel_btn, 2, 1)
        header_layout.addWidget(self.export_pdf_btn, 2, 2)
        header_layout.addWidget(self.export_trend_btn, 2, 3)

        header_layout.addWidget(self.start_btn, 1, 0)
        header_layout.addWidget(self.stop_btn, 1, 1)
        header_layout.addWidget(self.record_start_btn, 1, 2)
        header_layout.addWidget(self.record_stop_btn, 1, 3)
        header_layout.addWidget(self.status, 1, 4, 1, 2)

        header_layout.addWidget(self.rec_label, 0, 6)
        header_layout.addWidget(self.rec_time_label, 0, 7)
        header_layout.addWidget(self.rec_count_label, 1, 6)
        header_layout.addWidget(self.rec_file_label, 1, 7)
        header_layout.setColumnStretch(5, 1)

        main_layout.addWidget(header)

        self.tabs = QTabWidget()

        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Sensor", "Wert", "Einheit", "Status", "Alarm", "Min", "Max", "Zeit"]
        )
        self.table.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.table, "Dashboard")

        self.trend_plot = pg.PlotWidget()
        self.trend_plot.setBackground("w")
        self.trend_plot.showGrid(x=True, y=True)
        self.trend_plot.addLegend()
        self.trend_plot.setLabel("left", "Wert")
        self.trend_plot.setLabel("bottom", "Zeitpunkte")
        self.tabs.addTab(self.trend_plot, "Trend")

        self.telegram_log = QPlainTextEdit()
        self.telegram_log.setReadOnly(True)
        self.tabs.addTab(self.telegram_log, "Telegramme")

        self.decoder_log = QPlainTextEdit()
        self.decoder_log.setReadOnly(True)
        self.tabs.addTab(self.decoder_log, "Decoder")

        self.explorer_table = QTableWidget(0, 3)
        self.explorer_table.setHorizontalHeaderLabels(["Feld", "Wert", "Bedeutung"])
        self.explorer_table.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.explorer_table, "Telegramm-Explorer")

        config_widget = QWidget()
        config_layout = QVBoxLayout(config_widget)
        config_buttons = QHBoxLayout()
        self.config_save_btn = QPushButton("Sensor-Konfiguration speichern")
        self.config_reload_btn = QPushButton("Konfiguration neu laden")
        config_buttons.addWidget(self.config_save_btn)
        config_buttons.addWidget(self.config_reload_btn)
        config_buttons.addStretch()
        config_layout.addLayout(config_buttons)
        self.config_table = QTableWidget(0, 7)
        self.config_table.setHorizontalHeaderLabels(["ID", "Name", "Unit-Code", "Warn Low", "Warn High", "Alarm Low", "Alarm High"])
        self.config_table.horizontalHeader().setStretchLastSection(True)
        config_layout.addWidget(self.config_table)
        self.tabs.addTab(config_widget, "Sensor-Konfiguration")

        self.alarm_log = QPlainTextEdit()
        self.alarm_log.setReadOnly(True)
        self.tabs.addTab(self.alarm_log, "Alarme")

        self.export_log = QPlainTextEdit()
        self.export_log.setReadOnly(True)
        self.tabs.addTab(self.export_log, "Export")

        self.diagnostics = QPlainTextEdit()
        self.diagnostics.setReadOnly(True)
        self.tabs.addTab(self.diagnostics, "Diagnose")

        main_layout.addWidget(self.tabs)
        self.setCentralWidget(root)

        self.refresh_btn.clicked.connect(self.refresh_ports)
        self.start_btn.clicked.connect(self.start)
        self.stop_btn.clicked.connect(self.stop)
        self.record_start_btn.clicked.connect(self.start_recording)
        self.record_stop_btn.clicked.connect(self.stop_recording)
        self.player_load_btn.clicked.connect(self.load_player_file)
        self.export_snapshot_btn.clicked.connect(self.export_csv)
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.export_pdf_btn.clicked.connect(self.export_pdf_report)
        self.export_trend_btn.clicked.connect(self.export_trend_png)
        self.config_save_btn.clicked.connect(self.save_sensor_config_from_table)
        self.config_reload_btn.clicked.connect(self.reload_sensor_config)
        self.populate_config_table()

    def load_sensor_config_from_file(self):
        self.config_data = ensure_sensor_config()
        self.sensor_names.clear()
        self.sensor_unit_codes.clear()
        self.limits.clear()
        for entry in self.config_data.get("sensors", []):
            sid = int(entry.get("id", 0))
            if sid <= 0:
                continue
            self.sensor_names[sid] = str(entry.get("name", f"Sensor {sid}"))
            self.sensor_unit_codes[sid] = int(entry.get("unit_code", 0))
            self.limits[sid] = {
                "warn_low": float(entry.get("warn_low", 0)),
                "warn_high": float(entry.get("warn_high", 0)),
                "alarm_low": float(entry.get("alarm_low", 0)),
                "alarm_high": float(entry.get("alarm_high", 0)),
            }

    def populate_config_table(self):
        sensors = self.config_data.get("sensors", [])
        self.config_table.setRowCount(len(sensors))
        columns = ["id", "name", "unit_code", "warn_low", "warn_high", "alarm_low", "alarm_high"]
        for row, entry in enumerate(sensors):
            for col, key in enumerate(columns):
                self.config_table.setItem(row, col, QTableWidgetItem(str(entry.get(key, ""))))

    def save_sensor_config_from_table(self):
        sensors = []
        for row in range(self.config_table.rowCount()):
            try:
                entry = {
                    "id": int(self.config_table.item(row, 0).text()),
                    "name": self.config_table.item(row, 1).text(),
                    "unit_code": int(self.config_table.item(row, 2).text()),
                    "warn_low": float(self.config_table.item(row, 3).text().replace(",", ".")),
                    "warn_high": float(self.config_table.item(row, 4).text().replace(",", ".")),
                    "alarm_low": float(self.config_table.item(row, 5).text().replace(",", ".")),
                    "alarm_high": float(self.config_table.item(row, 6).text().replace(",", ".")),
                }
                sensors.append(entry)
            except Exception as exc:
                self.log(f"Konfigurationsfehler in Zeile {row + 1}: {exc}")
                return
        self.config_data = {"sensors": sensors}
        CONFIG_FILE.write_text(json.dumps(self.config_data, indent=2, ensure_ascii=False), encoding="utf-8")
        self.load_sensor_config_from_file()
        self.sensor_values.clear()
        self.active_alarms.clear()
        self.table.setRowCount(0)
        self.alarm_log.appendPlainText(f"{datetime.now():%H:%M:%S} [INFO] Sensor-Konfiguration gespeichert")
        self.log("Sensor-Konfiguration gespeichert")

    def reload_sensor_config(self):
        self.load_sensor_config_from_file()
        self.populate_config_table()
        self.sensor_values.clear()
        self.active_alarms.clear()
        self.table.setRowCount(0)
        self.log("Sensor-Konfiguration neu geladen")

    def refresh_ports(self):
        self.port_box.clear()

        if list_ports is None:
            self.port_box.addItem("pyserial nicht verfügbar", "")
            return

        ports = list(list_ports.comports())

        if not ports:
            self.port_box.addItem("Kein COM-Port gefunden", "")
            return

        for port in ports:
            self.port_box.addItem(f"{port.device} - {port.description}", port.device)

    def start(self):
        self.stop(clear_status=False)
        self.telegram_log.clear()
        self.decoder_log.clear()
        self.explorer_table.setRowCount(0)
        self.buffer.clear()

        mode = self.mode.currentText()

        if mode == "Simulator":
            self.status.setText("Status: Simulator läuft")
            self.log("Simulator gestartet")
            self.sim_timer.start(250)
            return

        if mode == "RS422":
            self.start_rs422()
            return

        if mode == "Player":
            self.start_player()
            return

    def stop(self, clear_status: bool = True):
        self.sim_timer.stop()
        self.serial_timer.stop()
        self.player_timer.stop()

        if self.serial_port is not None:
            try:
                self.serial_port.close()
                self.log("RS422 geschlossen")
            except Exception as exc:
                self.log(f"Fehler beim Schließen: {exc}")
            self.serial_port = None

        if clear_status:
            self.status.setText("Status: gestoppt")

    def start_rs422(self):
        if serial is None:
            self.status.setText("Status: pyserial fehlt")
            self.log("pyserial ist nicht verfügbar")
            return

        port = self.port_box.currentData()

        if not port:
            self.status.setText("Status: Kein COM-Port")
            self.log("Kein COM-Port ausgewählt")
            return

        try:
            self.serial_port = serial.Serial(
                port=port,
                baudrate=38400,
                bytesize=8,
                parity=serial.PARITY_ODD,
                stopbits=1,
                timeout=0,
            )

            self.status.setText(f"Status: RS422 verbunden {port}")
            self.log(f"RS422 geöffnet: {port}, 38400 Baud, 8O1")
            self.serial_timer.start(50)

        except Exception as exc:
            self.status.setText("Status: RS422 Fehler")
            self.log(f"RS422 Fehler: {exc}")

    def read_serial(self):
        if self.serial_port is None:
            return

        try:
            waiting = self.serial_port.in_waiting

            if waiting <= 0:
                return

            data = self.serial_port.read(waiting)
            self.byte_count += len(data)

            if data:
                self.telegram_log.appendPlainText(
                    f"RX {datetime.now():%H:%M:%S.%f}  {hex_string(data)}"
                )
                self.process_bytes(data)

        except Exception as exc:
            self.log(f"Lesefehler RS422: {exc}")
            self.stop()

    def update_simulator(self):
        self.t += 0.25

        oil_temp = 82 + 3 * math.sin(self.t)
        oil_pressure = 5.2 + 0.2 * math.sin(self.t / 2)
        rpm = 1480 + 60 * math.sin(self.t / 3)
        coolant = 79 + 2.5 * math.cos(self.t)
        exhaust = 460 + random.randint(-8, 8)

        if int(self.t) % 30 > 24:
            oil_temp += 14
        if int(self.t) % 40 > 34:
            exhaust += 120

        packets = bytearray()
        packets += make_packet(1, self.sensor_unit_codes.get(1, 1), int(oil_temp * 10))
        packets += make_packet(2, self.sensor_unit_codes.get(2, 2), int(oil_pressure * 100))
        packets += make_packet(3, self.sensor_unit_codes.get(3, 3), int(rpm))
        packets += make_packet(4, self.sensor_unit_codes.get(4, 1), int(coolant * 10))
        packets += make_packet(5, self.sensor_unit_codes.get(5, 1), int(exhaust * 10))

        self.byte_count += len(packets)
        self.telegram_log.appendPlainText(f"RX {hex_string(bytes(packets))}")
        self.process_bytes(bytes(packets))

    def process_bytes(self, data: bytes):
        self.buffer.extend(data)

        while True:
            if SYNC_BYTE not in self.buffer:
                self.buffer.clear()
                return

            sync_index = self.buffer.index(SYNC_BYTE)

            if sync_index > 0:
                del self.buffer[:sync_index]

            if len(self.buffer) < PACKET_LENGTH:
                return

            packet = bytes(self.buffer[:PACKET_LENGTH])
            del self.buffer[:PACKET_LENGTH]

            self.handle_packet(packet)

    def handle_packet(self, packet: bytes):
        self.decode_packet(packet)
        self.record_packet(packet)

    def decode_packet(self, packet: bytes):
        if len(packet) != PACKET_LENGTH or packet[0] != SYNC_BYTE:
            return

        header = packet[1]
        destination = packet[2]
        source = packet[3]
        channel = packet[4]
        unit_code = packet[5] & 0x0F

        # Byte 7: Bit 6 wird hier als Sensorfehlerbit ausgewertet.
        # Die unteren 6 Bit plus Byte 8 bilden aktuell den 14-Bit-Rohwert.
        raw_msb_byte = packet[6]
        sensor_fault = bool(raw_msb_byte & 0x40)
        msb = raw_msb_byte & 0x3F
        lsb = packet[7]

        raw_value = (msb << 8) | lsb
        unit = self.units.get(unit_code, "")

        if unit == "°C":
            value = raw_value / 10.0
        elif unit in {"bar", "V", "A", "%"}:
            value = raw_value / 100.0
        else:
            value = float(raw_value)

        name = self.sensor_names.get(channel, f"Sensor {channel}")
        now = datetime.now().strftime("%H:%M:%S")

        state, alarm_text = self.evaluate_sensor(channel, name, value, unit, sensor_fault)

        old = self.sensor_values.get(channel)
        min_value = value if old is None else min(old["min"], value)
        max_value = value if old is None else max(old["max"], value)

        self.sensor_values[channel] = {
            "id": channel,
            "name": name,
            "value": value,
            "unit": unit,
            "state": state,
            "alarm": alarm_text,
            "min": min_value,
            "max": max_value,
            "time": now,
            "sensor_fault": sensor_fault,
        }

        self.packet_count += 1

        self.decoder_log.appendPlainText(
            f"{now}  CH={channel}  {name} = {value:.2f} {unit}  "
            f"STATUS={state} SENSORFEHLER={sensor_fault} "
            f"SRC={source} DST={destination} HEADER={header:02X} WT={header & 0x0F}"
        )

        self.update_explorer(
            packet=packet,
            header=header,
            destination=destination,
            source=source,
            channel=channel,
            unit_code=unit_code,
            raw_msb_byte=raw_msb_byte,
            msb=msb,
            lsb=lsb,
            raw_value=raw_value,
            value=value,
            unit=unit,
            name=name,
            state=state,
            sensor_fault=sensor_fault,
        )

        self.update_dashboard()
        self.update_trend(channel, value, name)

    def update_explorer(
        self,
        packet: bytes,
        header: int,
        destination: int,
        source: int,
        channel: int,
        unit_code: int,
        raw_msb_byte: int,
        msb: int,
        lsb: int,
        raw_value: int,
        value: float,
        unit: str,
        name: str,
        state: str,
        sensor_fault: bool,
    ):
        word_type = header & 0x0F
        word_type_text = self.word_types.get(word_type, "unbekannter Word Type")
        rows = [
            ("Telegramm", hex_string(packet), "Vollständiges 8-Byte-Telegramm"),
            ("Byte 1 / Sync", f"0x{packet[0]:02X}", "Synchronisationsbyte"),
            ("Byte 2 / Header", f"0x{header:02X}", f"Word Type {word_type}: {word_type_text}"),
            ("Byte 3 / Destination", str(destination), "Zieladresse"),
            ("Byte 4 / Source", str(source), "Senderadresse"),
            ("Byte 5 / Channel", str(channel), name),
            ("Byte 6 / Unit", str(unit_code), unit if unit else "unbekannte Einheit"),
            ("Byte 7 / MSB", f"0x{raw_msb_byte:02X}", f"Wert-MSB={msb}, Sensorfehlerbit={sensor_fault}"),
            ("Byte 8 / LSB", f"0x{lsb:02X}", "Wert-LSB"),
            ("Raw Value", str(raw_value), "Rohwert vor Skalierung"),
            ("Skalierter Wert", f"{value:.2f} {unit}", name),
            ("Sensorfehler", "JA" if sensor_fault else "Nein", "Auswertung Bit 6 in Byte 7"),
            ("Status", state, "Ergebnis aus Sensorfehler und Grenzwertprüfung"),
        ]

        self.explorer_table.setRowCount(len(rows))
        for row, (field, value_text, meaning) in enumerate(rows):
            self.explorer_table.setItem(row, 0, QTableWidgetItem(field))
            self.explorer_table.setItem(row, 1, QTableWidgetItem(value_text))
            self.explorer_table.setItem(row, 2, QTableWidgetItem(meaning))

    def evaluate_sensor(self, channel: int, name: str, value: float, unit: str, sensor_fault: bool = False):
        limits = self.limits.get(channel)

        alarm_key = channel
        text = ""

        if sensor_fault:
            state = "SENSORFEHLER"
            text = f"{name}: Sensorfehlerbit gesetzt"
        elif not limits:
            state = "OK"
        elif value <= limits["alarm_low"] or value >= limits["alarm_high"]:
            state = "ALARM"
            text = f"{name}: {value:.2f} {unit} außerhalb Alarmgrenze"
        elif value <= limits["warn_low"] or value >= limits["warn_high"]:
            state = "WARNUNG"
            text = f"{name}: {value:.2f} {unit} außerhalb Warngrenze"
        else:
            state = "OK"

        if state in {"WARNUNG", "ALARM", "SENSORFEHLER"}:
            old = self.active_alarms.get(alarm_key)
            if old != text:
                self.active_alarms[alarm_key] = text
                self.alarm_log.appendPlainText(f"{datetime.now():%H:%M:%S} [{state}] {text}")
        else:
            if alarm_key in self.active_alarms:
                self.alarm_log.appendPlainText(
                    f"{datetime.now():%H:%M:%S} [OK] {name}: Wert wieder normal"
                )
                del self.active_alarms[alarm_key]

        return state, text

    def update_dashboard(self):
        sensors = list(self.sensor_values.values())
        sensors.sort(key=lambda s: s["id"])

        self.table.setRowCount(len(sensors))

        for row, sensor in enumerate(sensors):
            values = [
                str(sensor["id"]),
                sensor["name"],
                f"{sensor['value']:.2f}",
                sensor["unit"],
                sensor["state"],
                sensor["alarm"],
                f"{sensor['min']:.2f}",
                f"{sensor['max']:.2f}",
                sensor["time"],
            ]

            for col, text in enumerate(values):
                item = QTableWidgetItem(text)

                if col in (0, 2, 6, 7):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

                if sensor["state"] in {"ALARM", "SENSORFEHLER"}:
                    item.setBackground(Qt.red)
                    item.setForeground(Qt.white)
                elif sensor["state"] == "WARNUNG":
                    item.setBackground(Qt.yellow)
                    item.setForeground(Qt.black)
                else:
                    item.setBackground(Qt.white)
                    item.setForeground(Qt.black)

                self.table.setItem(row, col, item)

    def update_trend(self, channel: int, value: float, name: str):
        self.trend_data[channel].append(value)

        if channel not in self.curves:
            pen = pg.mkPen(width=2)
            self.curves[channel] = self.trend_plot.plot([], [], pen=pen, name=name)

        y_values = list(self.trend_data[channel])
        x_values = list(range(len(y_values)))
        self.curves[channel].setData(x_values, y_values)

    def start_recording(self):
        if self.recording:
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.record_file_path = REC_DIR / f"{timestamp}.mcslog"
        self.record_file = self.record_file_path.open("w", encoding="utf-8")
        self.recording = True
        self.record_start_time = datetime.now()
        self.record_packet_count = 0
        self.log(f"Aufnahme gestartet: {self.record_file_path}")
        self.update_recording_status(force=True)

    def stop_recording(self):
        if not self.recording:
            return

        if self.record_file:
            self.record_file.close()
        self.record_file = None
        self.recording = False
        self.log("Aufnahme gestoppt")
        self.update_recording_status(force=True)

    def record_packet(self, packet: bytes):
        if not self.recording or self.record_file is None:
            return

        self.record_file.write(f"{datetime.now().isoformat()} {packet.hex().upper()}\n")
        self.record_file.flush()
        self.record_packet_count += 1

    def update_recording_status(self, force: bool = False):
        if self.recording:
            self.rec_blink = not self.rec_blink
            lamp = "🔴" if self.rec_blink else "⚫"
            self.rec_label.setText(f"{lamp} REC")
            self.rec_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #b00020;")

            elapsed = datetime.now() - self.record_start_time if self.record_start_time else None
            if elapsed:
                total_seconds = int(elapsed.total_seconds())
                h = total_seconds // 3600
                m = (total_seconds % 3600) // 60
                s = total_seconds % 60
                self.rec_time_label.setText(f"{h:02d}:{m:02d}:{s:02d}")

            self.rec_count_label.setText(f"Telegramme: {self.record_packet_count}")
            if self.record_file_path:
                self.rec_file_label.setText(f"Datei: {self.record_file_path.name}")
        else:
            self.rec_label.setText("⚫ REC AUS")
            self.rec_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #333;")
            self.rec_time_label.setText("00:00:00")
            self.rec_count_label.setText(f"Telegramme: {self.record_packet_count}")
            if self.record_file_path:
                self.rec_file_label.setText(f"Datei: {self.record_file_path.name}")
            else:
                self.rec_file_label.setText("Datei: -")


    def _sensor_rows_for_export(self):
        sensors = list(self.sensor_values.values())
        sensors.sort(key=lambda s: s["id"])
        rows = []
        for sensor in sensors:
            rows.append({
                "id": sensor["id"],
                "name": sensor["name"],
                "value": sensor["value"],
                "unit": sensor["unit"],
                "state": sensor["state"],
                "alarm": sensor["alarm"],
                "min": sensor["min"],
                "max": sensor["max"],
                "time": sensor["time"],
            })
        return rows

    def export_csv(self):
        rows = self._sensor_rows_for_export()
        if not rows:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} Keine Sensordaten zum Export vorhanden")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"sensor_snapshot_{timestamp}.csv"
        fields = ["id", "name", "value", "unit", "state", "alarm", "min", "max", "time"]

        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, delimiter=";")
            writer.writeheader()
            writer.writerows(rows)

        self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} CSV exportiert: {path}")
        self.log(f"CSV exportiert: {path}")

    def _table_widget_rows(self, table: QTableWidget):
        headers = []
        for col in range(table.columnCount()):
            header_item = table.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else f"Spalte {col + 1}")

        rows = []
        for row in range(table.rowCount()):
            row_values = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                row_values.append(item.text() if item else "")
            rows.append(row_values)
        return headers, rows

    def _plain_text_rows(self, text_widget: QPlainTextEdit, title: str):
        lines = text_widget.toPlainText().splitlines()
        if not lines:
            lines = ["Keine Einträge vorhanden"]
        return [title], [[line] for line in lines]

    def _autosize_sheet(self, ws):
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    def _write_sheet(self, wb, sheet_name: str, headers: list, rows: list, title: str | None = None):
        ws = wb.create_sheet(sheet_name)

        current_row = 1
        if title:
            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=1).font = self._excel_title_font
            current_row += 2

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self._excel_header_font
            cell.fill = self._excel_header_fill
            cell.border = self._excel_border

        for r_index, row_values in enumerate(rows, start=current_row + 1):
            for c_index, value in enumerate(row_values, start=1):
                cell = ws.cell(row=r_index, column=c_index, value=value)
                cell.border = self._excel_border

        ws.freeze_panes = ws.cell(row=current_row + 1, column=1)
        self._autosize_sheet(ws)
        return ws

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as exc:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} Excel-Export nicht möglich: {exc}")
            self.log(f"Excel-Export nicht möglich: {exc}")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"mcs4_complete_export_{timestamp}.xlsx"

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        self._excel_title_font = Font(bold=True, size=14)
        self._excel_header_font = Font(bold=True, color="FFFFFF")
        self._excel_header_fill = PatternFill("solid", fgColor="1F4E78")
        thin = Side(style="thin", color="D9E2F3")
        self._excel_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Übersicht
        ws = wb.create_sheet("Übersicht")
        overview = [
            ("MCS-4 Monitor Export", ""),
            ("Erstellt", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Software-Version", APP_VERSION),
            ("Modus", self.mode.currentText()),
            ("COM-Port", str(self.port_box.currentData())),
            ("Telegramme dekodiert", self.packet_count),
            ("Bytes empfangen/erzeugt", self.byte_count),
            ("Sensoren im Dashboard", len(self.sensor_values)),
            ("Aktive Alarme", len(self.active_alarms)),
            ("Aufnahme aktiv", "Ja" if self.recording else "Nein"),
            ("Aufnahme-Telegramme", self.record_packet_count),
            ("Aufnahmedatei", self.record_file_path.name if self.record_file_path else "-"),
            ("Player-Datei", self.player_file_path.name if self.player_file_path else "-"),
        ]
        for row, (key, value) in enumerate(overview, start=1):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            if row == 1:
                ws.cell(row=row, column=1).font = self._excel_title_font
            else:
                ws.cell(row=row, column=1).font = self._excel_header_font
                ws.cell(row=row, column=1).fill = self._excel_header_fill
        self._autosize_sheet(ws)

        # Dashboard
        headers, rows = self._table_widget_rows(self.table)
        self._write_sheet(wb, "Dashboard", headers, rows, "Aktuelle Sensorwerte")

        # Alarme
        alarm_headers, alarm_rows = self._plain_text_rows(self.alarm_log, "Alarmmeldungen")
        self._write_sheet(wb, "Alarme", alarm_headers, alarm_rows, "Alarmhistorie")

        # Telegramme
        telegram_headers, telegram_rows = self._plain_text_rows(self.telegram_log, "Telegramme")
        self._write_sheet(wb, "Telegramme", telegram_headers, telegram_rows, "Telegrammmonitor")

        # Decoder
        decoder_headers, decoder_rows = self._plain_text_rows(self.decoder_log, "Decoder")
        self._write_sheet(wb, "Decoder", decoder_headers, decoder_rows, "Decoder-Ausgabe")

        # Diagnose
        diag_headers, diag_rows = self._plain_text_rows(self.diagnostics, "Diagnose")
        self._write_sheet(wb, "Diagnose", diag_headers, diag_rows, "Diagnoseinformationen")

        # Telegramm-Explorer
        headers, rows = self._table_widget_rows(self.explorer_table)
        self._write_sheet(wb, "Telegramm-Explorer", headers, rows, "Letztes Telegramm")

        # Sensor-Konfiguration
        headers, rows = self._table_widget_rows(self.config_table)
        self._write_sheet(wb, "Sensor-Konfiguration", headers, rows, "Aktuelle Sensorkonfiguration")

        # Export-Protokoll
        export_headers, export_rows = self._plain_text_rows(self.export_log, "Export-Protokoll")
        self._write_sheet(wb, "Export-Protokoll", export_headers, export_rows, "Export-Protokoll")

        # Trends als Datenreihen
        ws_trend = wb.create_sheet("Trenddaten")
        ws_trend.cell(row=1, column=1, value="Index")
        ws_trend.cell(row=1, column=1).font = self._excel_header_font
        ws_trend.cell(row=1, column=1).fill = self._excel_header_fill

        channels = sorted(self.trend_data.keys())
        for col, channel in enumerate(channels, start=2):
            name = self.sensor_names.get(channel, f"Sensor {channel}")
            unit = self.sensor_values.get(channel, {}).get("unit", "")
            header = f"{name} [{unit}]" if unit else name
            cell = ws_trend.cell(row=1, column=col, value=header)
            cell.font = self._excel_header_font
            cell.fill = self._excel_header_fill

        max_len = max((len(values) for values in self.trend_data.values()), default=0)
        for idx in range(max_len):
            ws_trend.cell(row=idx + 2, column=1, value=idx + 1)
            for col, channel in enumerate(channels, start=2):
                values = list(self.trend_data[channel])
                if idx < len(values):
                    ws_trend.cell(row=idx + 2, column=col, value=values[idx])
        self._autosize_sheet(ws_trend)

        wb.save(path)
        self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} Komplett-Excel exportiert: {path}")
        self.log(f"Komplett-Excel exportiert: {path}")


    def export_trend_png(self):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"trend_{timestamp}.png"
        pixmap = self.trend_plot.grab()
        if pixmap.save(str(path)):
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} Trend PNG exportiert: {path}")
            self.log(f"Trend PNG exportiert: {path}")
        else:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} Trend PNG Export fehlgeschlagen")

    def export_pdf_report(self):
        rows = self._sensor_rows_for_export()
        if not rows:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} Keine Sensordaten für PDF-Bericht vorhanden")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"mcs4_report_{timestamp}.pdf"

        alarm_items = list(self.active_alarms.values())
        alarm_html = "".join(f"<li>{alarm}</li>" for alarm in alarm_items) or "<li>Keine aktiven Alarme</li>"

        sensor_rows = ""
        for sensor in rows:
            sensor_rows += (
                "<tr>"
                f"<td>{sensor['id']}</td>"
                f"<td>{sensor['name']}</td>"
                f"<td>{sensor['value']:.2f}</td>"
                f"<td>{sensor['unit']}</td>"
                f"<td>{sensor['state']}</td>"
                f"<td>{sensor['alarm']}</td>"
                f"<td>{sensor['min']:.2f}</td>"
                f"<td>{sensor['max']:.2f}</td>"
                f"<td>{sensor['time']}</td>"
                "</tr>"
            )

        html = f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 10pt; }}
                h1 {{ color: #1f2937; }}
                h2 {{ color: #374151; margin-top: 18px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #999; padding: 4px; }}
                th {{ background: #e5e7eb; }}
                .meta {{ color: #555; }}
            </style>
        </head>
        <body>
            <h1>MCS-4 Diagnosebericht</h1>
            <p class="meta">Erstellt: {datetime.now():%Y-%m-%d %H:%M:%S}</p>
            <p class="meta">Software-Version: {APP_VERSION}</p>
            <p class="meta">Modus: {self.mode.currentText()} | COM-Port: {self.port_box.currentData()}</p>
            <h2>Zusammenfassung</h2>
            <ul>
                <li>Dekodierte Telegramme: {self.packet_count}</li>
                <li>Bytes empfangen/erzeugt: {self.byte_count}</li>
                <li>Sensoren im Dashboard: {len(self.sensor_values)}</li>
                <li>Aktive Alarme: {len(self.active_alarms)}</li>
            </ul>
            <h2>Sensorübersicht</h2>
            <table>
                <tr>
                    <th>ID</th><th>Sensor</th><th>Wert</th><th>Einheit</th><th>Status</th>
                    <th>Alarm</th><th>Min</th><th>Max</th><th>Zeit</th>
                </tr>
                {sensor_rows}
            </table>
            <h2>Aktive Alarme</h2>
            <ul>{alarm_html}</ul>
            <h2>Hinweis</h2>
            <p>Der MCS-4 Decoder ist aktuell ein erweiterbarer Basisdecoder. Die vollständige Word-Type- und Skalierungslogik wird schrittweise anhand der MCS-4 Dokumentation vervollständigt.</p>
        </body>
        </html>
        """

        document = QTextDocument()
        document.setHtml(html)
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(str(path))
        document.print_(printer)

        self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} PDF-Bericht exportiert: {path}")
        self.log(f"PDF-Bericht exportiert: {path}")

    def load_player_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "MCS-Logdatei laden",
            str(REC_DIR),
            "MCS Log (*.mcslog);;Alle Dateien (*.*)",
        )
        if not path:
            return

        self.player_file_path = Path(path)
        self.player_packets.clear()
        self.player_index = 0

        with self.player_file_path.open("r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 2:
                    hex_part = parts[-1]
                    try:
                        self.player_packets.append(bytes.fromhex(hex_part))
                    except ValueError:
                        pass

        self.mode.setCurrentText("Player")
        self.status.setText(f"Status: Player-Datei geladen ({len(self.player_packets)} Telegramme)")
        self.log(f"Player-Datei geladen: {self.player_file_path}")

    def start_player(self):
        if not self.player_packets:
            self.status.setText("Status: Keine Player-Datei geladen")
            self.log("Keine Player-Datei geladen")
            return

        self.player_index = 0
        self.status.setText("Status: Player läuft")
        self.log("Player gestartet")
        self.player_timer.start(100)

    def player_tick(self):
        if self.player_index >= len(self.player_packets):
            self.player_timer.stop()
            self.status.setText("Status: Player fertig")
            self.log("Player fertig")
            return

        packet = self.player_packets[self.player_index]
        self.player_index += 1
        self.byte_count += len(packet)
        self.telegram_log.appendPlainText(f"PLAY {hex_string(packet)}")
        self.process_bytes(packet)

    def update_diagnostics(self):
        text = (
            f"Zeit: {datetime.now():%Y-%m-%d %H:%M:%S}\n"
            f"Modus: {self.mode.currentText()}\n"
            f"Telegramme dekodiert: {self.packet_count}\n"
            f"Bytes empfangen/erzeugt: {self.byte_count}\n"
            f"Aktueller COM-Port: {self.port_box.currentData()}\n"
            f"RS422 aktiv: {'Ja' if self.serial_port else 'Nein'}\n"
            f"Sensoren im Dashboard: {len(self.sensor_values)}\n"
            f"Aktive Alarme: {len(self.active_alarms)}\n"
            f"Aufnahme aktiv: {'Ja' if self.recording else 'Nein'}\n"
            f"Aufnahme-Telegramme: {self.record_packet_count}\n"
            f"Player-Datei: {self.player_file_path.name if self.player_file_path else '-'}\n"
            f"Konfigurationsdatei: {CONFIG_FILE}\n"
            f"Exportordner: {EXPORT_DIR}\n"
            f"Decoder: MCS-4 Basisdecoder, WordType/Einheiten erweiterbar\n"
        )
        self.diagnostics.setPlainText(text)

    def log(self, text):
        self.telegram_log.appendPlainText(f"LOG {datetime.now():%H:%M:%S} {text}")

    def closeEvent(self, event):
        self.stop_recording()
        self.stop()
        event.accept()


def main():
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

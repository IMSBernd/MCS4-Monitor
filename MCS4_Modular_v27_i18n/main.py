import csv
import json
import math
import random
import sys
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path

import pyqtgraph as pg
from PySide6.QtCore import QTimer, Qt
from PySide6.QtGui import QTextDocument
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

try:
    import serial
    from serial.tools import list_ports
except Exception:
    serial = None
    list_ports = None

from license_system import LicenseManager, LicenseStatus, machine_id


APP_VERSION = "2.4"
SYNC_BYTE = 0xFF
PACKET_LENGTH = 8

WORD_LENGTHS = {0: 8, 1: 8, 2: 5, 3: 5, 4: 5, 5: 5, 6: 8, 7: 8}


def decode_word_type(header: int) -> int:
    """Byte 2 bits D3..D0 define the word type (PDF appendix 1)."""
    return header & 0x0F


def decode_word_flags(header: int) -> int:
    """Byte 2 bits D7..D4 are word-specific extensions/flags."""
    return (header >> 4) & 0x0F


REC_DIR = Path("recordings")
REC_DIR.mkdir(exist_ok=True)
EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)
CONFIG_FILE = Path("sensor_config.json")
LANG_DIR = Path("lang")
LANG_DIR.mkdir(exist_ok=True)


def load_language_file(language_code: str) -> dict:
    path = LANG_DIR / f"{language_code}.json"
    if not path.exists():
        path = LANG_DIR / "en.json"
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def hex_string(data: bytes) -> str:
    return " ".join(f"{b:02X}" for b in data)


def encode_data_packet(measuring_point: int, page: int, line: int, raw_value: int, *, source: int = 2, destination: int = 1, sensor_fault: bool = False, negative: bool = False) -> bytes:
    """Create a protocol-like MCS-4 data value telegram for simulator tests.

    According to Appendix 1/2, byte 7 contains sign, sensor-fault and the
    upper 5 bits of the 12-bit value; byte 8 contains the lower 7 bits.
    """
    raw_value = max(0, min(int(raw_value), 0x0FFF))
    byte6 = ((page & 0x1F) << 3) | (line & 0x07)
    byte7 = ((0x40 if sensor_fault else 0) | (0x20 if negative else 0) | ((raw_value >> 7) & 0x1F))
    byte8 = raw_value & 0x7F
    return bytes([
        SYNC_BYTE,
        0x00,
        destination & 0x7F,
        source & 0x3F,
        measuring_point & 0x7F,
        byte6 & 0xFE if byte6 == 0xFF else byte6,
        byte7 & 0x7F,
        byte8 & 0x7F,
    ])


def raw_from_scaled(value: float, page: int, line: int) -> int:
    info = SCALING_TABLE.get((page, line))
    if not info or not info.get("range") or not info.get("factor"):
        return int(value)
    return int(round(float(value) * float(info["factor"]) / float(info["range"])))


def make_packet(channel: int, unit_code: int, value: int) -> bytes:
    # Backward-compatible simulator helper. unit_code is treated as a legacy unit.
    legacy = {1: (3, 0), 2: (1, 1), 3: (8, 0)}
    page, line = legacy.get(unit_code, (0, 0))
    return encode_data_packet(channel, page, line, value)

# Extracted from Appendix 12 of the MCS-4 data format documentation.
# Key is (PAGE, LINE). The value is scaled as: physical = raw * range / factor.
# This replaces the old incorrect interpretation of byte 6 as a simple unit code.
SCALING_TABLE = {
    (0, 0): {"unit": "bar", "range": 1.000, "factor": 3200},
    (0, 1): {"unit": "bar", "range": 1.000, "factor": 2000},
    (0, 2): {"unit": "bar", "range": 1.750, "factor": 3500},
    (0, 3): {"unit": "bar", "range": 2.000, "factor": 2000},
    (0, 4): {"unit": "bar", "range": 2.800, "factor": 2800},
    (1, 0): {"unit": "bar", "range": 4.00, "factor": 3200},
    (1, 1): {"unit": "bar", "range": 7.00, "factor": 2800},
    (1, 2): {"unit": "bar", "range": 8.00, "factor": 3200},
    (1, 3): {"unit": "bar", "range": 10.00, "factor": 2000},
    (1, 4): {"unit": "bar", "range": 16.00, "factor": 3200},
    (1, 5): {"unit": "bar", "range": 20.00, "factor": 2000},
    (1, 6): {"unit": "bar", "range": 30.00, "factor": 3000},
    (1, 7): {"unit": "bar", "range": 32.00, "factor": 3200},
    (2, 0): {"unit": "bar", "range": 50.0, "factor": 2000},
    (2, 1): {"unit": "bar", "range": 60.0, "factor": 2560},
    (2, 2): {"unit": "bar", "range": 70.0, "factor": 2800},
    (2, 3): {"unit": "bar", "range": 100.0, "factor": 2000},
    (2, 4): {"unit": "bar", "range": 128.0, "factor": 2560},
    (2, 5): {"unit": "bar", "range": 250.0, "factor": 2500},
    (2, 6): {"unit": "bar", "range": 300.0, "factor": 3000},
    (2, 7): {"unit": "bar", "range": 400.0, "factor": 4000},
    (3, 0): {"unit": "°C", "range": 200.0, "factor": 2000},
    (4, 0): {"unit": "°C", "range": 1000.0, "factor": 2000},
    (5, 0): {"unit": "%", "range": 120.0, "factor": 2400},
    (5, 1): {"unit": "%", "range": 100.0, "factor": 2000},
    (6, 0): {"unit": "mm", "range": 20.00, "factor": 2000},
    (6, 1): {"unit": "mm", "range": 20.00, "factor": 4000},
    (6, 2): {"unit": "mm", "range": 40.00, "factor": 4000},
    (7, 0): {"unit": "deg", "range": 60.0, "factor": 2400},
    (7, 1): {"unit": "deg", "range": 50.0, "factor": 2000},
    (7, 2): {"unit": "deg", "range": 120.0, "factor": 2400},
    (7, 3): {"unit": "deg", "range": 100.0, "factor": 4000},
    (8, 0): {"unit": "rpm", "range": 4000.0, "factor": 4000},
    (9, 0): {"unit": "m3", "range": 1.00, "factor": 1600},
    (9, 1): {"unit": "m3", "range": 5.00, "factor": 2000},
    (9, 2): {"unit": "m3", "range": 10.00, "factor": 2000},
    (9, 3): {"unit": "m3", "range": 15.00, "factor": 3000},
    (9, 4): {"unit": "m3", "range": 25.00, "factor": 2500},
    (9, 5): {"unit": "m3", "range": 35.00, "factor": 3500},
    (10, 0): {"unit": "pts", "range": 60.0, "factor": 2400},
    (11, 0): {"unit": "V", "range": 250.0, "factor": 2000},
    (11, 1): {"unit": "V", "range": 400.0, "factor": 3200},
    (11, 2): {"unit": "V", "range": 500.0, "factor": 2000},
    (11, 3): {"unit": "V", "range": 150.0, "factor": 2400},
    (12, 0): {"unit": "A", "range": 10.0, "factor": 1600},
    (12, 1): {"unit": "A", "range": 15.0, "factor": 2400},
    (12, 2): {"unit": "A", "range": 25.0, "factor": 2000},
    (12, 3): {"unit": "A", "range": 200.0, "factor": 2000},
    (12, 4): {"unit": "A", "range": 400.0, "factor": 4000},
    (13, 0): {"unit": "kW", "range": 50.0, "factor": 2000},
    (13, 1): {"unit": "kW", "range": 100.0, "factor": 2000},
    (13, 2): {"unit": "kW", "range": 150.0, "factor": 3000},
    (13, 3): {"unit": "kW", "range": 200.0, "factor": 2000},
    (13, 4): {"unit": "kW", "range": 250.0, "factor": 2500},
    (13, 5): {"unit": "kW", "range": 300.0, "factor": 3000},
    (13, 6): {"unit": "kW", "range": 400.0, "factor": 4000},
    (16, 0): {"unit": "t", "range": 1.00, "factor": 1600},
    (16, 1): {"unit": "t", "range": 2.00, "factor": 3200},
    (16, 2): {"unit": "t", "range": 4.00, "factor": 3200},
    (16, 3): {"unit": "t", "range": 8.00, "factor": 3200},
    (16, 4): {"unit": "t", "range": 16.00, "factor": 3200},
    (16, 5): {"unit": "t", "range": 32.00, "factor": 3200},
    (18, 0): {"unit": "kW", "range": 1200.0, "factor": 2400},
    (19, 0): {"unit": "A", "range": 1000.0, "factor": 4000},
    (19, 1): {"unit": "A", "range": 1400.0, "factor": 2800},
    (19, 2): {"unit": "A", "range": 2400.0, "factor": 2400},
    (20, 0): {"unit": "kN", "range": 1400.0, "factor": 2800},
    (21, 0): {"unit": "krpm", "range": 40.00, "factor": 4000},
    (22, 0): {"unit": "krpm", "range": 400.0, "factor": 4000},
    (22, 1): {"unit": "krpm", "range": 80.0, "factor": 3200},
    (23, 0): {"unit": "knm", "range": 400.0, "factor": 4000},
    (24, 0): {"unit": "mils", "range": 10.00, "factor": 2000},
    (25, 0): {"unit": "pts", "range": 4000.0, "factor": 4000},
    (26, 0): {"unit": "V", "range": 10.00, "factor": 4000},
    (26, 1): {"unit": "V", "range": 16.00, "factor": 3200},
    (26, 2): {"unit": "V", "range": 40.00, "factor": 4000},
    (27, 0): {"unit": "s", "range": 400.0, "factor": 4000},
}


def decode_page_line(byte6: int) -> tuple[int, int]:
    return (byte6 >> 3) & 0x1F, byte6 & 0x07


def decode_12bit_value(byte7: int, byte8: int) -> tuple[int, bool, bool]:
    sensor_fault = bool(byte7 & 0x40)
    negative = bool(byte7 & 0x20)
    raw = ((byte7 & 0x1F) << 7) | (byte8 & 0x7F)
    return raw, negative, sensor_fault


def scale_raw_value(raw: int, page: int, line: int, negative: bool = False) -> tuple[float, str, str]:
    info = SCALING_TABLE.get((page, line))
    signed_raw = -raw if negative else raw
    if not info:
        return float(signed_raw), "", "keine Appendix-12-Zuordnung"
    value = float(signed_raw) * float(info["range"]) / float(info["factor"])
    return value, info["unit"], f"Page {page}, Line {line}: Bereich {info['range']} {info['unit']}, Faktor {info['factor']}"




def sensor_key(measuring_point: int, page: int, line: int) -> str:
    """Composite key for one logical sensor value.

    MCS-4 uses Byte 5 as measuring point and Byte 6 as Page/Line.
    The same measuring point can legitimately occur with different Page/Line
    combinations. Therefore the dashboard and configuration must not use the
    measuring point alone as unique sensor identifier.
    """
    return f"{int(measuring_point)}:{int(page)}:{int(line)}"


def sensor_key_text(measuring_point: int, page: int, line: int) -> str:
    return f"MP {int(measuring_point)} / P{int(page)} L{int(line)}"


def sort_sensor_id(value) -> tuple[int, int, int, str]:
    try:
        if isinstance(value, str) and ":" in value:
            a, b, c = value.split(":", 2)
            return int(a), int(b), int(c), value
        return int(value), 0, 0, str(value)
    except Exception:
        return 999999, 999999, 999999, str(value)

def default_sensor_config() -> dict:
    return {
        "sensors": [
            {"key": "1:3:0", "id": 1, "name": "Oil temperature", "page": 3, "line": 0, "unit": "°C", "warn_low": 0, "warn_high": 90, "alarm_low": -10, "alarm_high": 95},
            {"key": "2:1:1", "id": 2, "name": "Oil pressure", "page": 1, "line": 1, "unit": "bar", "warn_low": 4.5, "warn_high": 6.0, "alarm_low": 4.0, "alarm_high": 6.5},
            {"key": "3:8:0", "id": 3, "name": "Engine speed", "page": 8, "line": 0, "unit": "rpm", "warn_low": 500, "warn_high": 1900, "alarm_low": 300, "alarm_high": 2100},
            {"key": "4:3:0", "id": 4, "name": "Coolant temperature", "page": 3, "line": 0, "unit": "°C", "warn_low": 0, "warn_high": 88, "alarm_low": -10, "alarm_high": 95},
            {"key": "5:4:0", "id": 5, "name": "Exhaust temperature", "page": 4, "line": 0, "unit": "°C", "warn_low": 0, "warn_high": 520, "alarm_low": -10, "alarm_high": 560}
        ]
    }


def ensure_sensor_config() -> dict:
    if not CONFIG_FILE.exists():
        CONFIG_FILE.write_text(json.dumps(default_sensor_config(), indent=2, ensure_ascii=False), encoding="utf-8")
    with CONFIG_FILE.open("r", encoding="utf-8") as f:
        return json.load(f)


class MainWindow(QMainWindow):
    def __init__(self, license_status: LicenseStatus | None = None):
        super().__init__()
        self.license_status = license_status
        self.language = "en"
        self.translations = load_language_file(self.language)
        self.setWindowTitle(self.tr("window_title", version=APP_VERSION))
        self.resize(1360, 860)

        self.t = 0.0
        self.packet_count = 0
        self.byte_count = 0
        self.serial_port = None
        self.buffer = bytearray()

        self.recording = False
        self.record_start_time: datetime | None = None
        self.record_file = None
        self.record_file_path: Path | None = None
        self.record_packet_count = 0
        self.rec_blink = False

        self.player_packets: list[bytes] = []
        self.player_index = 0
        self.player_file_path: Path | None = None

        self.config_data = ensure_sensor_config()

        # Appendix-12-Skalierung wird über Byte 6 als Page/Line ausgewertet.
        # self.units bleibt nur für alte Konfigurations-/Simulatorfunktionen erhalten.
        self.units = {0: "", 1: "°C", 2: "bar", 3: "rpm", 4: "V", 5: "A", 6: "%"}

        self.sensor_names = {}
        self.sensor_unit_codes = {}
        self.limits = {}
        self.load_sensor_config_from_file()

        # WordType laut PDF: Byte 2, Bits D3..D0. Bits D7..D4 sind word-spezifische Flags.
        self.word_types = {
            0: "Data Value / Messwert",
            1: "Limit Value / Grenzwert",
            2: "Alarm Message / Alarmmeldung",
            3: "Control Command / Steuerbefehl",
            4: "Binary Signal",
            5: "Key Identification / Kennung",
            6: "Status Message",
            7: "Curve Transfer / Kurve",
            8: "nicht definiert / reserviert",
            9: "nicht definiert / reserviert",
            10: "nicht definiert / reserviert",
            11: "nicht definiert / reserviert",
            12: "nicht definiert / reserviert",
            13: "nicht definiert / reserviert",
            14: "nicht definiert / reserviert",
        }

        self.sensor_values = {}
        self.active_alarms = {}

        # Analyzer/Lernmodus: erfasst alle empfangenen MCS-4 Data-Value-Kombinationen
        # als eindeutigen Schlüssel Measuring Point + Page + Line. Damit können echte
        # Anlagen-Measuring Pointe später gezielt benannt und konfiguriert werden.
        self.analyzer_seen = {}

        # Live Protocol Analyzer / Version 2.1
        self.protocol_wordtype_counts = defaultdict(int)
        self.protocol_mp_counts = defaultdict(int)
        self.protocol_key_counts = defaultdict(int)
        self.protocol_invalid_count = 0
        self.protocol_total_count = 0
        self.protocol_packet_times = deque(maxlen=1000)

        self.trend_data = defaultdict(lambda: deque(maxlen=250))
        self.curves = {}

        self._build_ui()
        self.refresh_ports()

        self.sim_timer = QTimer(self)
        self.sim_timer.timeout.connect(self.update_simulator)

        self.serial_timer = QTimer(self)
        self.serial_timer.timeout.connect(self.read_serial)

        self.player_timer = QTimer(self)
        self.player_timer.timeout.connect(self.player_tick)

        self.diag_timer = QTimer(self)
        self.diag_timer.timeout.connect(self.update_diagnostics)
        self.diag_timer.start(1000)

        self.rec_timer = QTimer(self)
        self.rec_timer.timeout.connect(self.update_recording_status)
        self.rec_timer.start(500)

    def tr(self, key: str, **kwargs) -> str:
        value = self.translations.get(key, key)
        if isinstance(value, list):
            return value
        try:
            return str(value).format(**kwargs)
        except Exception:
            return str(value)

    def change_language(self) -> None:
        code = self.language_box.currentData() if hasattr(self, "language_box") else "en"
        self.language = code or "en"
        self.translations = load_language_file(self.language)
        self.apply_language()

    def apply_language(self) -> None:
        self.setWindowTitle(self.tr("window_title", version=APP_VERSION))
        self.title_label.setText(self.tr("app_title"))
        self.version_label.setText(f"Version {APP_VERSION}")
        self.language_label.setText(self.tr("language"))
        self.mode_label.setText(self.tr("mode"))
        self.port_label.setText(self.tr("com_port"))
        self.refresh_btn.setText(self.tr("refresh_ports"))
        self.start_btn.setText(self.tr("start"))
        self.stop_btn.setText(self.tr("stop"))
        self.record_start_btn.setText(self.tr("start_recording"))
        self.record_stop_btn.setText(self.tr("stop_recording"))
        self.player_load_btn.setText(self.tr("load_player_file"))
        self.export_snapshot_btn.setText(self.tr("export_csv"))
        self.export_excel_btn.setText(self.tr("export_excel"))
        self.export_pdf_btn.setText(self.tr("export_pdf"))
        self.export_trend_btn.setText(self.tr("export_trend_png"))
        if not self.serial_port and not self.sim_timer.isActive() and not self.player_timer.isActive():
            self.status.setText(self.tr("status_ready"))
        if not self.recording:
            self.rec_label.setText(self.tr("rec_off"))
            self.rec_file_label.setText(self.tr("file_none"))
        self.rec_count_label.setText(self.tr("telegrams_count", count=self.record_packet_count))
        self.update_license_label()
        self.tabs.setTabText(0, self.tr("tabs_dashboard"))
        self.tabs.setTabText(1, self.tr("tabs_trend"))
        self.tabs.setTabText(2, self.tr("tabs_telegrams"))
        self.tabs.setTabText(3, self.tr("tabs_decoder"))
        self.tabs.setTabText(4, self.tr("tabs_telegram_explorer"))
        self.tabs.setTabText(5, self.tr("tabs_mcs4_analyzer"))
        self.tabs.setTabText(6, self.tr("tabs_protocol_statistics"))
        self.tabs.setTabText(7, self.tr("tabs_sensor_configuration"))
        self.tabs.setTabText(8, self.tr("tabs_alarms"))
        self.tabs.setTabText(9, self.tr("tabs_export"))
        self.tabs.setTabText(10, self.tr("tabs_diagnostics"))
        self.table.setHorizontalHeaderLabels(self.translations.get("dashboard_headers", []))
        self.explorer_table.setHorizontalHeaderLabels(self.translations.get("explorer_headers", []))
        self.analyzer_table.setHorizontalHeaderLabels(self.translations.get("analyzer_headers", []))
        self.config_table.setHorizontalHeaderLabels(self.translations.get("config_headers", []))
        self.analyzer_hint.setText(self.tr("mcs4_analyzer_hint"))
        self.learn_save_btn.setText(self.tr("learning_mode_save_name"))
        self.learn_key_label.setText(self.tr("detected_key"))
        self.learn_name_label.setText(self.tr("name"))
        self.learn_name_edit.setPlaceholderText(self.tr("placeholder_sensor_name"))
        self.config_save_btn.setText(self.tr("save_sensor_configuration"))
        self.config_reload_btn.setText(self.tr("reload_configuration"))
        self.config_cleanup_btn.setText(self.tr("clean_unused_sensors"))
        self.trend_plot.setLabel("left", self.tr("value_axis")) if "value_axis" in self.translations else self.trend_plot.setLabel("left", "Value")

    def update_license_label(self) -> None:
        if self.license_status and self.license_status.valid:
            self.license_label.setText(self.tr("license_line", customer=self.license_status.customer, type=self.license_status.license_type, days=self.license_status.days_left))
        else:
            self.license_label.setText(self.tr("license_not_checked"))

    def _build_ui(self):
        root = QWidget()
        main_layout = QVBoxLayout(root)

        title_row = QHBoxLayout()
        self.title_label = QLabel(self.tr("app_title"))
        title = self.title_label
        title.setStyleSheet("font-size: 22px; font-weight: bold;")
        self.version_label = QLabel(f"Version {APP_VERSION}")
        version = self.version_label
        version.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.license_label = QLabel("")
        self.license_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.license_label.setStyleSheet("color: #555;")
        title_row.addWidget(title)
        title_row.addStretch()
        title_row.addWidget(self.license_label)
        title_row.addSpacing(16)
        title_row.addWidget(version)
        main_layout.addLayout(title_row)

        header = QFrame()
        header.setFrameShape(QFrame.StyledPanel)
        header.setStyleSheet("QFrame { background: #f4f6f8; border-radius: 6px; }")
        header_layout = QGridLayout(header)

        self.language_label = QLabel(self.tr("language"))
        self.language_box = QComboBox()
        self.language_box.addItem("English", "en")
        self.language_box.addItem("Deutsch", "de")
        self.mode_label = QLabel(self.tr("mode"))
        self.mode = QComboBox()
        self.mode.addItems(["Simulator", "RS422", "Player"])
        self.port_label = QLabel(self.tr("com_port"))
        self.port_box = QComboBox()
        self.refresh_btn = QPushButton(self.tr("refresh_ports"))
        self.start_btn = QPushButton(self.tr("start"))
        self.stop_btn = QPushButton(self.tr("stop"))
        self.record_start_btn = QPushButton(self.tr("start_recording"))
        self.record_stop_btn = QPushButton(self.tr("stop_recording"))
        self.player_load_btn = QPushButton(self.tr("load_player_file"))
        self.export_snapshot_btn = QPushButton(self.tr("export_csv"))
        self.export_excel_btn = QPushButton(self.tr("export_excel"))
        self.export_pdf_btn = QPushButton(self.tr("export_pdf"))
        self.export_trend_btn = QPushButton(self.tr("export_trend_png"))

        self.status = QLabel(self.tr("status_ready"))
        self.status.setStyleSheet("font-weight: bold;")

        self.rec_label = QLabel(self.tr("rec_off"))
        self.rec_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #333;")
        self.rec_time_label = QLabel("00:00:00")
        self.rec_count_label = QLabel(self.tr("telegrams_count", count=0))
        self.rec_file_label = QLabel(self.tr("file_none"))

        header_layout.addWidget(self.mode_label, 0, 0)
        header_layout.addWidget(self.mode, 0, 1)
        header_layout.addWidget(self.port_label, 0, 2)
        header_layout.addWidget(self.port_box, 0, 3)
        header_layout.addWidget(self.refresh_btn, 0, 4)
        header_layout.addWidget(self.player_load_btn, 0, 5)
        header_layout.addWidget(self.export_snapshot_btn, 2, 0)
        header_layout.addWidget(self.export_excel_btn, 2, 1)
        header_layout.addWidget(self.export_pdf_btn, 2, 2)
        header_layout.addWidget(self.export_trend_btn, 2, 3)

        header_layout.addWidget(self.start_btn, 1, 0)
        header_layout.addWidget(self.stop_btn, 1, 1)
        header_layout.addWidget(self.record_start_btn, 1, 2)
        header_layout.addWidget(self.record_stop_btn, 1, 3)
        header_layout.addWidget(self.status, 1, 4, 1, 2)

        header_layout.addWidget(self.rec_label, 0, 6)
        header_layout.addWidget(self.rec_time_label, 0, 7)
        header_layout.addWidget(self.rec_count_label, 1, 6)
        header_layout.addWidget(self.rec_file_label, 1, 7)
        header_layout.addWidget(self.language_label, 2, 6)
        header_layout.addWidget(self.language_box, 2, 7)
        header_layout.setColumnStretch(5, 1)

        main_layout.addWidget(header)

        self.tabs = QTabWidget()

        self.table = QTableWidget(0, 11)
        self.table.setHorizontalHeaderLabels(self.translations.get("dashboard_headers", []))
        self.table.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.table, self.tr("tabs_dashboard"))

        self.trend_plot = pg.PlotWidget()
        self.trend_plot.setBackground("w")
        self.trend_plot.showGrid(x=True, y=True)
        self.trend_plot.addLegend()
        self.trend_plot.setLabel("left", "Value")
        self.trend_plot.setLabel("bottom", "Zeitpunkte")
        self.tabs.addTab(self.trend_plot, self.tr("tabs_trend"))

        self.telegram_log = QPlainTextEdit()
        self.telegram_log.setReadOnly(True)
        self.tabs.addTab(self.telegram_log, self.tr("tabs_telegrams"))

        self.decoder_log = QPlainTextEdit()
        self.decoder_log.setReadOnly(True)
        self.tabs.addTab(self.decoder_log, self.tr("tabs_decoder"))

        self.explorer_table = QTableWidget(0, 3)
        self.explorer_table.setHorizontalHeaderLabels(self.translations.get("explorer_headers", []))
        self.explorer_table.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.explorer_table, self.tr("tabs_telegram_explorer"))

        analyzer_widget = QWidget()
        analyzer_layout = QVBoxLayout(analyzer_widget)
        self.analyzer_hint = QLabel(self.tr("mcs4_analyzer_hint"))
        analyzer_hint = self.analyzer_hint
        analyzer_layout.addWidget(analyzer_hint)

        learn_row = QHBoxLayout()
        self.learn_key_box = QComboBox()
        self.learn_name_edit = QLineEdit()
        self.learn_name_edit.setPlaceholderText(self.tr("placeholder_sensor_name"))
        self.learn_save_btn = QPushButton(self.tr("learning_mode_save_name"))
        self.learn_key_label = QLabel(self.tr("detected_key"))
        self.learn_name_label = QLabel(self.tr("name"))
        learn_row.addWidget(self.learn_key_label)
        learn_row.addWidget(self.learn_key_box)
        learn_row.addWidget(self.learn_name_label)
        learn_row.addWidget(self.learn_name_edit)
        learn_row.addWidget(self.learn_save_btn)
        analyzer_layout.addLayout(learn_row)

        self.analyzer_table = QTableWidget(0, 13)
        self.analyzer_table.setHorizontalHeaderLabels(self.translations.get("analyzer_headers", []))
        self.analyzer_table.horizontalHeader().setStretchLastSection(True)
        analyzer_layout.addWidget(self.analyzer_table)
        self.tabs.addTab(analyzer_widget, self.tr("tabs_mcs4_analyzer"))

        self.protocol_stats_log = QPlainTextEdit()
        self.protocol_stats_log.setReadOnly(True)
        self.tabs.addTab(self.protocol_stats_log, self.tr("tabs_protocol_statistics"))

        config_widget = QWidget()
        config_layout = QVBoxLayout(config_widget)
        config_buttons = QHBoxLayout()
        self.config_save_btn = QPushButton(self.tr("save_sensor_configuration"))
        self.config_reload_btn = QPushButton(self.tr("reload_configuration"))
        self.config_cleanup_btn = QPushButton(self.tr("clean_unused_sensors"))
        config_buttons.addWidget(self.config_save_btn)
        config_buttons.addWidget(self.config_reload_btn)
        config_buttons.addWidget(self.config_cleanup_btn)
        config_buttons.addStretch()
        config_layout.addLayout(config_buttons)
        self.config_table = QTableWidget(0, 12)
        self.config_table.setHorizontalHeaderLabels(self.translations.get("config_headers", []))
        self.config_table.horizontalHeader().setStretchLastSection(True)
        config_layout.addWidget(self.config_table)
        self.tabs.addTab(config_widget, self.tr("tabs_sensor_configuration"))

        self.alarm_log = QPlainTextEdit()
        self.alarm_log.setReadOnly(True)
        self.tabs.addTab(self.alarm_log, self.tr("tabs_alarms"))

        self.export_log = QPlainTextEdit()
        self.export_log.setReadOnly(True)
        self.tabs.addTab(self.export_log, self.tr("tabs_export"))

        self.diagnostics = QPlainTextEdit()
        self.diagnostics.setReadOnly(True)
        self.tabs.addTab(self.diagnostics, self.tr("tabs_diagnostics"))

        main_layout.addWidget(self.tabs)
        self.setCentralWidget(root)

        self.language_box.currentIndexChanged.connect(self.change_language)
        self.refresh_btn.clicked.connect(self.refresh_ports)
        self.start_btn.clicked.connect(self.start)
        self.stop_btn.clicked.connect(self.stop)
        self.record_start_btn.clicked.connect(self.start_recording)
        self.record_stop_btn.clicked.connect(self.stop_recording)
        self.player_load_btn.clicked.connect(self.load_player_file)
        self.export_snapshot_btn.clicked.connect(self.export_csv)
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.export_pdf_btn.clicked.connect(self.export_pdf_report)
        self.export_trend_btn.clicked.connect(self.export_trend_png)
        self.config_save_btn.clicked.connect(self.save_sensor_config_from_table)
        self.config_reload_btn.clicked.connect(self.reload_sensor_config)
        self.config_cleanup_btn.clicked.connect(self.cleanup_unused_sensor_config)
        self.learn_save_btn.clicked.connect(self.save_learned_sensor_name)
        self.populate_config_table()

    def load_sensor_config_from_file(self):
        self.config_data = ensure_sensor_config()
        self.sensor_names.clear()
        self.sensor_unit_codes.clear()
        self.limits.clear()
        for entry in self.config_data.get("sensors", []):
            sid = int(entry.get("id", 0))
            if sid < 0:
                continue
            page = entry.get("page", "")
            line = entry.get("line", "")
            if (page == "" or line == "") and "unit_code" in entry:
                legacy = {1: (3, 0), 2: (1, 1), 3: (8, 0)}
                page, line = legacy.get(int(entry.get("unit_code", 0)), (0, 0))
            try:
                page = int(page)
                line = int(line)
            except Exception:
                page, line = 0, 0
            key = str(entry.get("key") or sensor_key(sid, page, line))
            self.sensor_names[key] = str(entry.get("name", f"Measuring Point {sid}"))
            self.sensor_unit_codes[key] = int(entry.get("unit_code", 0) or 0)
            self.limits[key] = {
                "warn_low": float(entry.get("warn_low", 0)),
                "warn_high": float(entry.get("warn_high", 0)),
                "alarm_low": float(entry.get("alarm_low", 0)),
                "alarm_high": float(entry.get("alarm_high", 0)),
            }

    def _config_entry_for_table(self, entry: dict) -> dict:
        sid = int(entry.get("id", 0))
        page = entry.get("page", "")
        line = entry.get("line", "")
        unit = entry.get("unit", "")
        # Migration alter Konfigurationen mit unit_code-Feld
        if (page == "" or line == "") and "unit_code" in entry:
            legacy = {1: (3, 0, "°C"), 2: (1, 1, "bar"), 3: (8, 0, "rpm")}
            page, line, unit = legacy.get(int(entry.get("unit_code", 0)), ("", "", unit))
        return {
            "key": str(entry.get("key") or sensor_key(sid, int(page or 0), int(line or 0))),
            "id": sid,
            "name": str(entry.get("name", f"Measuring Point {sid}")),
            "page": page,
            "line": line,
            "unit": unit,
            "warn_low": entry.get("warn_low", 0),
            "warn_high": entry.get("warn_high", 0),
            "alarm_low": entry.get("alarm_low", 0),
            "alarm_high": entry.get("alarm_high", 0),
        }

    def populate_config_table(self):
        sensors = [self._config_entry_for_table(e) for e in self.config_data.get("sensors", [])]
        sensors.sort(key=lambda e: (int(e.get("id", 0)), int(e.get("page", 0) or 0), int(e.get("line", 0) or 0)))
        self.config_table.setRowCount(len(sensors))
        columns = ["key", "id", "name", "page", "line", "unit", "warn_low", "warn_high", "alarm_low", "alarm_high"]
        active_keys = set(self.analyzer_seen.keys()) | set(self.sensor_values.keys())
        for row, entry in enumerate(sensors):
            entry_key = str(entry.get("key", ""))
            for col, key in enumerate(columns):
                self.config_table.setItem(row, col, QTableWidgetItem(str(entry.get(key, ""))))

            active = entry_key in active_keys
            active_item = QTableWidgetItem("Yes" if active else "No")
            status_item = QTableWidgetItem(self.tr("active_seen") if active else self.tr("inactive_seen"))
            if active:
                active_item.setBackground(Qt.green)
                status_item.setBackground(Qt.green)
            else:
                active_item.setBackground(Qt.lightGray)
                status_item.setBackground(Qt.lightGray)
            self.config_table.setItem(row, 10, active_item)
            self.config_table.setItem(row, 11, status_item)

    def save_sensor_config_from_table(self):
        sensors = []
        for row in range(self.config_table.rowCount()):
            try:
                def cell(col: int) -> str:
                    item = self.config_table.item(row, col)
                    return item.text().strip() if item else ""

                sid = int(cell(1))
                page = int(cell(3)) if cell(3) != "" else 0
                line = int(cell(4)) if cell(4) != "" else 0
                key = cell(0) or sensor_key(sid, page, line)
                entry = {
                    "key": key,
                    "id": sid,
                    "name": cell(2) or f"Measuring Point {sid}",
                    "page": page,
                    "line": line,
                    "unit": cell(5),
                    "warn_low": float(cell(6).replace(",", ".")) if cell(6) != "" else 0.0,
                    "warn_high": float(cell(7).replace(",", ".")) if cell(7) != "" else 0.0,
                    "alarm_low": float(cell(8).replace(",", ".")) if cell(8) != "" else 0.0,
                    "alarm_high": float(cell(9).replace(",", ".")) if cell(9) != "" else 0.0,
                }
                sensors.append(entry)
            except Exception as exc:
                self.log(f"Configuration error in row {row + 1}: {exc}")
                return
        self.config_data = {"sensors": sensors}
        CONFIG_FILE.write_text(json.dumps(self.config_data, indent=2, ensure_ascii=False), encoding="utf-8")
        self.load_sensor_config_from_file()
        self.sensor_values.clear()
        self.active_alarms.clear()
        self.table.setRowCount(0)
        self.alarm_log.appendPlainText(f"{datetime.now():%H:%M:%S} [INFO] {self.tr("config_saved")}")
        self.log("Sensor Configuration gespeichert")

    def reload_sensor_config(self):
        self.load_sensor_config_from_file()
        self.populate_config_table()
        self.sensor_values.clear()
        self.active_alarms.clear()
        self.table.setRowCount(0)
        self.log("Sensor Configuration neu geladen")

    def cleanup_unused_sensor_config(self):
        """Remove configuration entries that were not seen in the current session.

        This is useful after changing Page/Line decoding. Old entries such as
        5:3:0 remain in sensor_config.json until the user deliberately removes
        them. The cleanup keeps only keys that are currently present in the
        analyzer or dashboard.
        """
        active_keys = set(self.analyzer_seen.keys()) | set(self.sensor_values.keys())
        if not active_keys:
            self.log("Cleanup not possible: no sensors detected in this session yet")
            return

        sensors = list(self.config_data.get("sensors", []))
        keep = []
        removed = []
        for entry in sensors:
            key = str(entry.get("key") or sensor_key(int(entry.get("id", 0)), int(entry.get("page", 0) or 0), int(entry.get("line", 0) or 0)))
            if key in active_keys:
                keep.append(entry)
            else:
                removed.append(key)

        if not removed:
            self.log("Cleanup: no unused sensors found")
            self.populate_config_table()
            return

        answer = QMessageBox.question(
            self,
            "Sensor Configuration bereinigen",
            "Configuration entries not detected in the current session will be removed.\n\n"
            + "Entfernen: " + ", ".join(removed[:12])
            + (" ..." if len(removed) > 12 else "")
            + "\n\nFortfahren?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            self.log("Bereinigung abgebrochen")
            return

        self.config_data = {"sensors": keep}
        CONFIG_FILE.write_text(json.dumps(self.config_data, indent=2, ensure_ascii=False), encoding="utf-8")
        self.load_sensor_config_from_file()
        self.populate_config_table()
        self.log(f"Cleanup completed: {len(removed)} sensor(s) removed")
        self.alarm_log.appendPlainText(f"{datetime.now():%H:%M:%S} [INFO] {self.tr("config_cleaned", count=len(removed))}")

    def add_detected_sensor_to_config(self, measuring_point: int, page: int, line: int, unit: str) -> None:
        """Automatically add newly detected measuring points to the configuration table.

        The PDF defines byte 5 as measuring point number. The concrete sensor name
        is project-specific, therefore new points are stored as "Measuring Point n" and
        can be renamed by the user in the Sensor Configuration tab.
        """
        key = sensor_key(measuring_point, page, line)
        if key in self.sensor_names:
            return
        if measuring_point < 0 or measuring_point > 159:
            return

        entry = {
            "key": key,
            "id": measuring_point,
            "name": f"Measuring Point {measuring_point} P{page}L{line}",
            "page": page,
            "line": line,
            "unit": unit,
            "warn_low": 0.0,
            "warn_high": 0.0,
            "alarm_low": 0.0,
            "alarm_high": 0.0,
        }
        self.config_data.setdefault("sensors", []).append(entry)
        self.config_data["sensors"].sort(key=lambda e: (int(e.get("id", 0)), int(e.get("page", 0) or 0), int(e.get("line", 0) or 0)))
        CONFIG_FILE.write_text(json.dumps(self.config_data, indent=2, ensure_ascii=False), encoding="utf-8")
        self.load_sensor_config_from_file()
        self.populate_config_table()
        self.alarm_log.appendPlainText(
            f"{datetime.now():%H:%M:%S} [INFO] New sensor detected: {sensor_key_text(measuring_point, page, line)} ({unit or 'without unit'})"
        )

    def refresh_ports(self):
        self.port_box.clear()

        if list_ports is None:
            self.port_box.addItem("pyserial not available", "")
            return

        ports = list(list_ports.comports())

        if not ports:
            self.port_box.addItem("No COM port found", "")
            return

        for port in ports:
            self.port_box.addItem(f"{port.device} - {port.description}", port.device)

    def start(self):
        self.stop(clear_status=False)
        self.telegram_log.clear()
        self.decoder_log.clear()
        self.explorer_table.setRowCount(0)
        self.buffer.clear()

        mode = self.mode.currentText()

        if mode == "Simulator":
            self.status.setText(self.tr("status_simulator_running"))
            self.log("Simulator started")
            self.sim_timer.start(250)
            return

        if mode == "RS422":
            self.start_rs422()
            return

        if mode == "Player":
            self.start_player()
            return

    def stop(self, clear_status: bool = True):
        self.sim_timer.stop()
        self.serial_timer.stop()
        self.player_timer.stop()

        if self.serial_port is not None:
            try:
                self.serial_port.close()
                self.log(self.tr("log_rs422_closed"))
            except Exception as exc:
                self.log(f"Error while closing: {exc}")
            self.serial_port = None

        if clear_status:
            self.status.setText(self.tr("status_stopped"))

    def start_rs422(self):
        if serial is None:
            self.status.setText(self.tr("status_pyserial_missing"))
            self.log(self.tr("log_pyserial_missing"))
            return

        port = self.port_box.currentData()

        if not port:
            self.status.setText(self.tr("status_no_com_port"))
            self.log("No COM port selected")
            return

        try:
            self.serial_port = serial.Serial(
                port=port,
                baudrate=38400,
                bytesize=8,
                parity=serial.PARITY_ODD,
                stopbits=1,
                timeout=0,
            )

            self.status.setText(self.tr("status_rs422_connected", port=port))
            self.log(f"RS422 opened: {port}, 38400 Baud, 8O1")
            self.serial_timer.start(50)

        except Exception as exc:
            self.status.setText(self.tr("status_rs422_error"))
            self.log(self.tr("log_rs422_error", error=exc))

    def read_serial(self):
        if self.serial_port is None:
            return

        try:
            waiting = self.serial_port.in_waiting

            if waiting <= 0:
                return

            data = self.serial_port.read(waiting)
            self.byte_count += len(data)

            if data:
                self.telegram_log.appendPlainText(
                    f"RX {datetime.now():%H:%M:%S.%f}  {hex_string(data)}"
                )
                self.process_bytes(data)

        except Exception as exc:
            self.log(f"RS422 read error: {exc}")
            self.stop()

    def update_simulator(self):
        self.t += 0.25

        oil_temp = 82 + 3 * math.sin(self.t)
        oil_pressure = 5.2 + 0.2 * math.sin(self.t / 2)
        rpm = 1480 + 60 * math.sin(self.t / 3)
        coolant = 79 + 2.5 * math.cos(self.t)
        exhaust = 460 + random.randint(-8, 8)

        if int(self.t) % 30 > 24:
            oil_temp += 14
        if int(self.t) % 40 > 34:
            exhaust += 120

        packets = bytearray()
        # Simulator now uses the real MCS-4 byte-6 Page/Line principle.
        packets += encode_data_packet(1, 3, 0, raw_from_scaled(oil_temp, 3, 0))      # 000.0 °C
        packets += encode_data_packet(2, 1, 1, raw_from_scaled(oil_pressure, 1, 1))  # 00.00 bar
        packets += encode_data_packet(3, 8, 0, raw_from_scaled(rpm, 8, 0))           # 0000 rpm
        packets += encode_data_packet(4, 3, 0, raw_from_scaled(coolant, 3, 0))       # 000.0 °C
        packets += encode_data_packet(5, 4, 0, raw_from_scaled(exhaust, 4, 0))       # 0000 °C

        self.byte_count += len(packets)
        self.telegram_log.appendPlainText(f"RX {hex_string(bytes(packets))}")
        self.process_bytes(bytes(packets))

    def process_bytes(self, data: bytes):
        self.buffer.extend(data)

        while True:
            if SYNC_BYTE not in self.buffer:
                self.buffer.clear()
                return

            sync_index = self.buffer.index(SYNC_BYTE)
            if sync_index > 0:
                del self.buffer[:sync_index]

            if len(self.buffer) < 2:
                return

            header = self.buffer[1]
            word_type = decode_word_type(header)
            expected_length = WORD_LENGTHS.get(word_type)
            if expected_length is None:
                # Word types 8..14 are not defined in this document. We cannot know
                # their length safely, so discard only the sync byte and resync on
                # the next FF to avoid creating false measuring values.
                self.protocol_invalid_count += 1
                del self.buffer[0]
                continue

            if len(self.buffer) < expected_length:
                return

            packet = bytes(self.buffer[:expected_length])
            del self.buffer[:expected_length]
            self.handle_packet(packet)

    def handle_packet(self, packet: bytes):
        self.decode_packet(packet)

    def decode_packet(self, packet: bytes):
        if len(packet) < 5 or packet[0] != SYNC_BYTE:
            return

        header = packet[1]
        word_type = decode_word_type(header)
        word_type_text = self.word_types.get(word_type, "unbekannt")
        expected_length = WORD_LENGTHS.get(word_type, len(packet))
        destination = packet[2]
        source = packet[3]
        number = packet[4]
        now = datetime.now().strftime("%H:%M:%S")

        validation_errors = self.validate_packet(packet, word_type)
        self.register_protocol_packet(packet, word_type, number, None, None, None, validation_errors)

        if word_type != 0:
            self.packet_count += 1
            self.decoder_log.appendPlainText(
                f"{now}  WT={word_type} {word_type_text}  FLAGS=0x{decode_word_flags(header):X}  LEN={len(packet)}  "
                f"NO={number} SRC={source} DST={destination}  "
                f"{'INVALID: ' + '; '.join(validation_errors) if validation_errors else 'OK'}"
            )
            self.update_explorer_generic(packet, validation_errors)
            self.record_packet(packet)
            return

        if len(packet) != 8:
            self.decoder_log.appendPlainText(f"{now}  Data Value invalid: Length {len(packet)} instead of 8")
            self.update_explorer_generic(packet, [f"Data Value Length {len(packet)} instead of 8"])
            return

        measuring_point = number
        byte6 = packet[5]
        page, line = decode_page_line(byte6)
        raw_value, negative, sensor_fault = decode_12bit_value(packet[6], packet[7])
        value, unit, scaling_text = scale_raw_value(raw_value, page, line, negative)
        self.register_protocol_data_key(measuring_point, page, line, unit, raw_value, value)

        # Neu erkannte Measuring Pointe werden automatisch im Konfigurator angelegt.
        # Der Name kann anschließend manuell geändert und gespeichert werden.
        if not validation_errors:
            self.add_detected_sensor_to_config(measuring_point, page, line, unit)

        key = sensor_key(measuring_point, page, line)
        name = self.sensor_names.get(key, f"Measuring Point {measuring_point} P{page}L{line}")

        self.register_analyzer_value(
            key=key,
            measuring_point=measuring_point,
            page=page,
            line=line,
            name=name,
            unit=unit,
            value=value,
            raw_value=raw_value,
            packet=packet,
            status="INVALID" if validation_errors else "OK",
        )

        if validation_errors:
            self.packet_count += 1
            self.decoder_log.appendPlainText(
                f"{now}  WT=0 Data Value INVALID  MP={measuring_point}  "
                f"PAGE={page} LINE={line} RAW={raw_value}  Fehler: {'; '.join(validation_errors)}"
            )
            self.update_explorer_data(
                packet, header, destination, source, measuring_point, page, line,
                raw_value, negative, sensor_fault, value, unit, scaling_text, name,
                "INVALID", "; ".join(validation_errors)
            )
            self.record_packet(packet)
            return

        state, alarm_text = self.evaluate_sensor(key, name, value, unit, sensor_fault)

        old = self.sensor_values.get(key)
        min_value = value if old is None else min(old["min"], value)
        max_value = value if old is None else max(old["max"], value)

        self.sensor_values[key] = {
            "key": key,
            "id": measuring_point,
            "name": name,
            "value": value,
            "unit": unit,
            "state": state,
            "alarm": alarm_text,
            "min": min_value,
            "max": max_value,
            "time": now,
            "sensor_fault": sensor_fault,
            "page": page,
            "line": line,
            "raw": raw_value,
        }

        self.packet_count += 1

        self.decoder_log.appendPlainText(
            f"{now}  WT=0 Data Value  FLAGS=0x{decode_word_flags(header):X}  MP={measuring_point}  {name} = {value:.3f} {unit}  "
            f"PAGE={page} LINE={line} RAW={raw_value} STATUS={state} "
            f"SENSORERROR={sensor_fault} SRC={source} DST={destination}"
        )

        self.update_explorer_data(
            packet, header, destination, source, measuring_point, page, line,
            raw_value, negative, sensor_fault, value, unit, scaling_text, name, state, alarm_text
        )

        self.update_dashboard()
        self.update_trend(key, value, name)
        self.record_packet(packet)

    def validate_packet(self, packet: bytes, word_type: int) -> list[str]:
        errors = []
        expected = WORD_LENGTHS.get(word_type)
        if expected is not None and len(packet) != expected:
            errors.append(f"Length {len(packet)} instead of {expected}")
        if word_type not in WORD_LENGTHS:
            errors.append(f"WordType {word_type} laut Dokument nicht implementiert")
        if len(packet) >= 2:
            flags = decode_word_flags(packet[1])
            # Appendix 2: Data Value has Byte 2 = 0000 0000.
            if word_type == 0 and flags != 0:
                errors.append(f"Data-Value-Flags nicht 0 (Flags={flags})")
        if len(packet) >= 3 and packet[2] > 127:
            errors.append(f"Zieladresse >127 ({packet[2]})")
        if len(packet) >= 4 and packet[3] > 63:
            errors.append(f"Senderadresse >63 ({packet[3]})")
        if len(packet) >= 5:
            if word_type == 0 and packet[4] > 159:
                errors.append(f"Measuring point outside Data Value range 0..159 ({packet[4]})")
            elif packet[4] == 255:
                errors.append("Number 255 is not permitted")
        if len(packet) == 8:
            # In the MCS-4 data format, byte 7 and byte 8 are 7-bit fields plus flags.
            if packet[6] & 0x80:
                errors.append("Byte 7 D7 gesetzt, laut Format nicht erwartet")
            if packet[7] & 0x80:
                errors.append("Byte 8 D7 gesetzt, laut Format nicht erwartet")
        return errors

    def update_explorer_generic(self, packet: bytes, validation_errors: list[str]):
        header = packet[1] if len(packet) > 1 else 0
        word_type = decode_word_type(header)
        rows = [
            ("Telegram", hex_string(packet), "Received Telegram"),
            ("Length", str(len(packet)), f"Expected: {WORD_LENGTHS.get(word_type, '?')} Byte"),
            ("Byte 1 / Sync", f"0x{packet[0]:02X}" if packet else "-", "Synchronisationsbyte"),
            ("Byte 2 / Header", f"0x{header:02X}", f"Word Type {word_type}: {self.word_types.get(word_type, 'unbekannt')}"),
            ("Byte 2 Flags", f"0x{decode_word_flags(header):X}", "Bits D7..D4, wort-spezifische Erweiterungen"),
            ("Byte 3 / Destination", str(packet[2]) if len(packet) > 2 else "-", "Zieladresse"),
            ("Byte 4 / Source", str(packet[3]) if len(packet) > 3 else "-", "Senderadresse"),
            ("Byte 5 / Nummer", str(packet[4]) if len(packet) > 4 else "-", "Measuring Point / Alarmnummer / Befehlscode"),
            ("Validation", "OK" if not validation_errors else "ERROR", "; ".join(validation_errors) if validation_errors else "Telegram formal plausibel"),
        ]
        self._fill_explorer(rows)

    def update_explorer_data(
        self,
        packet: bytes,
        header: int,
        destination: int,
        source: int,
        measuring_point: int,
        page: int,
        line: int,
        raw_value: int,
        negative: bool,
        sensor_fault: bool,
        value: float,
        unit: str,
        scaling_text: str,
        name: str,
        state: str,
        alarm_text: str,
    ):
        word_type = decode_word_type(header)
        rows = [
            ("Telegram", hex_string(packet), "Vollständiges Telegram"),
            ("Length", str(len(packet)), "Data Value must have 8 bytes"),
            ("Byte 1 / Sync", f"0x{packet[0]:02X}", "Synchronisationsbyte"),
            ("Byte 2 / Header", f"0x{header:02X}", f"Word Type {word_type}: {self.word_types.get(word_type, 'unbekannt')}"),
            ("Byte 2 Flags", f"0x{decode_word_flags(header):X}", "Bits D7..D4, wort-spezifische Erweiterungen"),
            ("Byte 3 / Destination", str(destination), "Zieladresse 0..127"),
            ("Byte 4 / Source", str(source), "Senderadresse 0..63"),
            ("Byte 5 / Measuring Point", str(measuring_point), name),
            ("Byte 6", f"0x{packet[5]:02X}", "PAGE + LINE, nicht direkte Unit"),
            ("Page", str(page), "Appendix 12 page for unit/range"),
            ("Line", str(line), "Appendix-12-Zeile"),
            ("Byte 7", f"0x{packet[6]:02X}", "D6 Sensor fault, D5 Vorzeichen, D4..D0 Value"),
            ("Byte 8", f"0x{packet[7]:02X}", "D6..D0 Value"),
            ("Rohwert 12 Bit", str(raw_value), "((Byte7 & 0x1F) << 7) | (Byte8 & 0x7F)"),
            ("Vorzeichen", "negativ" if negative else "positiv", "Byte 7 D5"),
            ("Sensor fault", "YES" if sensor_fault else "No", "Byte 7 D6"),
            ("Skalierung", scaling_text, "Appendix 12"),
            ("Skalierter Value", f"{value:.3f} {unit}", name),
            ("Status", state, alarm_text if alarm_text else "OK"),
        ]
        self._fill_explorer(rows)

    def _fill_explorer(self, rows):
        self.explorer_table.setRowCount(len(rows))
        for row, (field, value_text, meaning) in enumerate(rows):
            self.explorer_table.setItem(row, 0, QTableWidgetItem(field))
            self.explorer_table.setItem(row, 1, QTableWidgetItem(value_text))
            self.explorer_table.setItem(row, 2, QTableWidgetItem(meaning))

    def register_analyzer_value(
        self,
        *,
        key: str,
        measuring_point: int,
        page: int,
        line: int,
        name: str,
        unit: str,
        value: float,
        raw_value: int,
        packet: bytes,
        status: str,
    ) -> None:
        old = self.analyzer_seen.get(key)
        if old is None:
            self.analyzer_seen[key] = {
                "key": key,
                "mp": measuring_point,
                "page": page,
                "line": line,
                "name": name,
                "unit": unit,
                "value": value,
                "raw": raw_value,
                "min": value,
                "max": value,
                "count": 1,
                "packet": hex_string(packet),
                "status": status,
            }
        else:
            old["name"] = name
            old["unit"] = unit
            old["value"] = value
            old["raw"] = raw_value
            old["min"] = min(float(old.get("min", value)), value)
            old["max"] = max(float(old.get("max", value)), value)
            old["count"] = int(old.get("count", 0)) + 1
            old["packet"] = hex_string(packet)
            old["status"] = status

        self.update_analyzer_table()
        # Aktiv-Status in Sensor Configuration aktualisieren
        if hasattr(self, "config_table"):
            self.populate_config_table()

    def update_analyzer_table(self) -> None:
        rows = list(self.analyzer_seen.values())
        rows.sort(key=lambda r: sort_sensor_id(r["key"]))
        self.analyzer_table.setRowCount(len(rows))

        existing_combo_keys = {self.learn_key_box.itemData(i) for i in range(self.learn_key_box.count())}

        for row, item in enumerate(rows):
            values = [
                item["key"],
                str(item["mp"]),
                str(item["page"]),
                str(item["line"]),
                item["name"],
                item["unit"],
                f"{float(item['value']):.3f}",
                str(item["raw"]),
                f"{float(item['min']):.3f}",
                f"{float(item['max']):.3f}",
                str(item["count"]),
                item["packet"],
                item["status"],
            ]
            for col, text in enumerate(values):
                cell = QTableWidgetItem(text)
                if item["status"] != "OK":
                    cell.setBackground(Qt.yellow)
                self.analyzer_table.setItem(row, col, cell)

            if item["key"] not in existing_combo_keys:
                self.learn_key_box.addItem(
                    f"{item['key']} - MP {item['mp']} P{item['page']} L{item['line']} - {item['name']}",
                    item["key"],
                )
                existing_combo_keys.add(item["key"])

    def save_learned_sensor_name(self) -> None:
        key = self.learn_key_box.currentData()
        name = self.learn_name_edit.text().strip()
        if not key:
            self.log("Learning mode: no sensor key selected")
            return
        if not name:
            self.log("Learning mode: please enter a sensor name")
            return

        found = False
        for entry in self.config_data.get("sensors", []):
            if str(entry.get("key")) == str(key):
                entry["name"] = name
                found = True
                break

        if not found:
            seen = self.analyzer_seen.get(key)
            if not seen:
                self.log(f"Lernmodus: Key {key} ist noch nicht im Analyzer vorhanden")
                return
            self.config_data.setdefault("sensors", []).append({
                "key": key,
                "id": int(seen["mp"]),
                "name": name,
                "page": int(seen["page"]),
                "line": int(seen["line"]),
                "unit": str(seen.get("unit", "")),
                "warn_low": 0.0,
                "warn_high": 0.0,
                "alarm_low": 0.0,
                "alarm_high": 0.0,
            })

        CONFIG_FILE.write_text(json.dumps(self.config_data, indent=2, ensure_ascii=False), encoding="utf-8")
        self.load_sensor_config_from_file()
        self.populate_config_table()
        if key in self.analyzer_seen:
            self.analyzer_seen[key]["name"] = name
            self.update_analyzer_table()
        self.log(f"Lernmodus: {key} als '{name}' gespeichert")
        self.alarm_log.appendPlainText(f"{datetime.now():%H:%M:%S} [INFO] {self.tr("learning_saved", key=key, name=name)}")

    def evaluate_sensor(self, sensor_id, name: str, value: float, unit: str, sensor_fault: bool = False):
        limits = self.limits.get(sensor_id)

        alarm_key = sensor_id
        text = ""

        if sensor_fault:
            state = "SENSORERROR"
            text = f"{name}: Sensor faultbit gesetzt"
        elif not limits:
            state = "OK"
        elif value <= limits["alarm_low"] or value >= limits["alarm_high"]:
            state = "ALARM"
            text = f"{name}: {value:.2f} {unit} outside alarm limit"
        elif value <= limits["warn_low"] or value >= limits["warn_high"]:
            state = "WARNUNG"
            text = f"{name}: {value:.2f} {unit} outside warning limit"
        else:
            state = "OK"

        if state in {"WARNUNG", "ALARM", "SENSORERROR"}:
            old = self.active_alarms.get(alarm_key)
            if old != text:
                self.active_alarms[alarm_key] = text
                self.alarm_log.appendPlainText(f"{datetime.now():%H:%M:%S} [{state}] {text}")
        else:
            if alarm_key in self.active_alarms:
                self.alarm_log.appendPlainText(
                    f"{datetime.now():%H:%M:%S} [OK] {name}: Value wieder normal"
                )
                del self.active_alarms[alarm_key]

        return state, text

    def update_dashboard(self):
        sensors = list(self.sensor_values.values())
        sensors.sort(key=lambda s: sort_sensor_id(s.get("key", s.get("id", 0))))

        self.table.setRowCount(len(sensors))

        for row, sensor in enumerate(sensors):
            values = [
                sensor.get("key", ""),
                str(sensor["id"]),
                str(sensor.get("page", "")),
                str(sensor.get("line", "")),
                sensor["name"],
                f"{sensor['value']:.2f}",
                sensor["unit"],
                sensor["state"],
                sensor["alarm"],
                f"{sensor['min']:.2f} / {sensor['max']:.2f}",
                sensor["time"],
            ]

            for col, text in enumerate(values):
                item = QTableWidgetItem(text)

                if col in (1, 2, 3, 5, 9):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

                if sensor["state"] in {"ALARM", "SENSORERROR"}:
                    item.setBackground(Qt.red)
                    item.setForeground(Qt.white)
                elif sensor["state"] == "WARNUNG":
                    item.setBackground(Qt.yellow)
                    item.setForeground(Qt.black)
                else:
                    item.setBackground(Qt.white)
                    item.setForeground(Qt.black)

                self.table.setItem(row, col, item)

    def update_trend(self, channel: int, value: float, name: str):
        self.trend_data[channel].append(value)

        if channel not in self.curves:
            pen = pg.mkPen(width=2)
            self.curves[channel] = self.trend_plot.plot([], [], pen=pen, name=name)

        y_values = list(self.trend_data[channel])
        x_values = list(range(len(y_values)))
        self.curves[channel].setData(x_values, y_values)

    def start_recording(self):
        if self.recording:
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.record_file_path = REC_DIR / f"{timestamp}.mcslog"
        self.record_file = self.record_file_path.open("w", encoding="utf-8")
        self.recording = True
        self.record_start_time = datetime.now()
        self.record_packet_count = 0
        self.log(f"Recording started: {self.record_file_path}")
        self.update_recording_status(force=True)

    def stop_recording(self):
        if not self.recording:
            return

        if self.record_file:
            self.record_file.close()
        self.record_file = None
        self.recording = False
        self.log("Recording stopped")
        self.update_recording_status(force=True)

    def record_packet(self, packet: bytes):
        if not self.recording or self.record_file is None:
            return

        self.record_file.write(f"{datetime.now().isoformat()} {packet.hex().upper()}\n")
        self.record_file.flush()
        self.record_packet_count += 1

    def update_recording_status(self, force: bool = False):
        if self.recording:
            self.rec_blink = not self.rec_blink
            lamp = "🔴" if self.rec_blink else "⚫"
            self.rec_label.setText(f"{lamp} REC")
            self.rec_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #b00020;")

            elapsed = datetime.now() - self.record_start_time if self.record_start_time else None
            if elapsed:
                total_seconds = int(elapsed.total_seconds())
                h = total_seconds // 3600
                m = (total_seconds % 3600) // 60
                s = total_seconds % 60
                self.rec_time_label.setText(f"{h:02d}:{m:02d}:{s:02d}")

            self.rec_count_label.setText(f"Telegrams: {self.record_packet_count}")
            if self.record_file_path:
                self.rec_file_label.setText(f"Datei: {self.record_file_path.name}")
        else:
            self.rec_label.setText("⚫ REC OFF")
            self.rec_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #333;")
            self.rec_time_label.setText("00:00:00")
            self.rec_count_label.setText(f"Telegrams: {self.record_packet_count}")
            if self.record_file_path:
                self.rec_file_label.setText(f"Datei: {self.record_file_path.name}")
            else:
                self.rec_file_label.setText("Datei: -")


    def _sensor_rows_for_export(self):
        sensors = list(self.sensor_values.values())
        sensors.sort(key=lambda s: sort_sensor_id(s.get("key", s.get("id", 0))))
        rows = []
        for sensor in sensors:
            rows.append({
                "key": sensor.get("key", ""),
                "id": sensor["id"],
                "page": sensor.get("page", ""),
                "line": sensor.get("line", ""),
                "name": sensor["name"],
                "value": sensor["value"],
                "unit": sensor["unit"],
                "state": sensor["state"],
                "alarm": sensor["alarm"],
                "min": sensor["min"],
                "max": sensor["max"],
                "time": sensor["time"],
            })
        return rows

    def export_csv(self):
        rows = self._sensor_rows_for_export()
        if not rows:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("no_sensor_data_export")}")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"sensor_snapshot_{timestamp}.csv"
        fields = ["id", "name", "value", "unit", "state", "alarm", "min", "max", "time"]

        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, delimiter=";")
            writer.writeheader()
            writer.writerows(rows)

        self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("csv_exported", path=path)}")
        self.log(f"CSV exportiert: {path}")

    def _table_widget_rows(self, table: QTableWidget):
        headers = []
        for col in range(table.columnCount()):
            header_item = table.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else f"Spalte {col + 1}")

        rows = []
        for row in range(table.rowCount()):
            row_values = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                row_values.append(item.text() if item else "")
            rows.append(row_values)
        return headers, rows

    def _plain_text_rows(self, text_widget: QPlainTextEdit, title: str):
        lines = text_widget.toPlainText().splitlines()
        if not lines:
            lines = ["No entries available"]
        return [title], [[line] for line in lines]

    def _autosize_sheet(self, ws):
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    def _write_sheet(self, wb, sheet_name: str, headers: list, rows: list, title: str | None = None):
        ws = wb.create_sheet(sheet_name)

        current_row = 1
        if title:
            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=1).font = self._excel_title_font
            current_row += 2

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self._excel_header_font
            cell.fill = self._excel_header_fill
            cell.border = self._excel_border

        for r_index, row_values in enumerate(rows, start=current_row + 1):
            for c_index, value in enumerate(row_values, start=1):
                cell = ws.cell(row=r_index, column=c_index, value=value)
                cell.border = self._excel_border

        ws.freeze_panes = ws.cell(row=current_row + 1, column=1)
        self._autosize_sheet(ws)
        return ws

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as exc:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("excel_failed", error=exc)}")
            self.log(f"Excel export not possible: {exc}")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"mcs4_complete_export_{timestamp}.xlsx"

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        self._excel_title_font = Font(bold=True, size=14)
        self._excel_header_font = Font(bold=True, color="FFFFFF")
        self._excel_header_fill = PatternFill("solid", fgColor="1F4E78")
        thin = Side(style="thin", color="D9E2F3")
        self._excel_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Overview
        ws = wb.create_sheet("Overview")
        overview = [
            ("MCS-4 Monitor Export", ""),
            ("Erstellt", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Software version", APP_VERSION),
            ("Modus", self.mode.currentText()),
            ("COM-Port", str(self.port_box.currentData())),
            ("Telegrams dekodiert", self.packet_count),
            ("Bytes empfangen/erzeugt", self.byte_count),
            ("Sensors in dashboard", len(self.sensor_values)),
            ("Aktive Alarms", len(self.active_alarms)),
            ("Recording active", "Yes" if self.recording else "No"),
            ("Aufnahme-Telegrams", self.record_packet_count),
            ("Recording file", self.record_file_path.name if self.record_file_path else "-"),
            ("Player-Datei", self.player_file_path.name if self.player_file_path else "-"),
        ]
        for row, (key, value) in enumerate(overview, start=1):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            if row == 1:
                ws.cell(row=row, column=1).font = self._excel_title_font
            else:
                ws.cell(row=row, column=1).font = self._excel_header_font
                ws.cell(row=row, column=1).fill = self._excel_header_fill
        self._autosize_sheet(ws)

        # Dashboard
        headers, rows = self._table_widget_rows(self.table)
        self._write_sheet(wb, "Dashboard", headers, rows, "Current sensor values")

        # Alarms
        alarm_headers, alarm_rows = self._plain_text_rows(self.alarm_log, "Alarmmeldungen")
        self._write_sheet(wb, "Alarms", alarm_headers, alarm_rows, "Alarm history")

        # Telegrams
        telegram_headers, telegram_rows = self._plain_text_rows(self.telegram_log, "Telegrams")
        self._write_sheet(wb, "Telegrams", telegram_headers, telegram_rows, "Telegrammonitor")

        # Decoder
        decoder_headers, decoder_rows = self._plain_text_rows(self.decoder_log, "Decoder")
        self._write_sheet(wb, "Decoder", decoder_headers, decoder_rows, "Decoder-Ausgabe")

        # Diagnostics
        diag_headers, diag_rows = self._plain_text_rows(self.diagnostics, "Diagnostics")
        self._write_sheet(wb, "Diagnostics", diag_headers, diag_rows, "Diagnosticsinformationen")

        # Telegram Explorer
        headers, rows = self._table_widget_rows(self.explorer_table)
        self._write_sheet(wb, "Telegram Explorer", headers, rows, "Letztes Telegram")

        # MCS-4 Analyzer
        headers, rows = self._table_widget_rows(self.analyzer_table)
        self._write_sheet(wb, "MCS-4 Analyzer", headers, rows, "Erkannte Measuring Pointe")

        # Sensor Configuration
        headers, rows = self._table_widget_rows(self.config_table)
        self._write_sheet(wb, "Sensor Configuration", headers, rows, "Current sensor configuration")

        # Export Log
        export_headers, export_rows = self._plain_text_rows(self.export_log, "Export Log")
        self._write_sheet(wb, "Export Log", export_headers, export_rows, "Export Log")

        # Trends als Datenreihen
        ws_trend = wb.create_sheet("Trenddaten")
        ws_trend.cell(row=1, column=1, value="Index")
        ws_trend.cell(row=1, column=1).font = self._excel_header_font
        ws_trend.cell(row=1, column=1).fill = self._excel_header_fill

        channels = sorted(self.trend_data.keys(), key=sort_sensor_id)
        for col, channel in enumerate(channels, start=2):
            name = self.sensor_values.get(channel, {}).get("name", self.sensor_names.get(channel, f"Sensor {channel}"))
            unit = self.sensor_values.get(channel, {}).get("unit", "")
            header = f"{name} [{unit}]" if unit else name
            cell = ws_trend.cell(row=1, column=col, value=header)
            cell.font = self._excel_header_font
            cell.fill = self._excel_header_fill

        max_len = max((len(values) for values in self.trend_data.values()), default=0)
        for idx in range(max_len):
            ws_trend.cell(row=idx + 2, column=1, value=idx + 1)
            for col, channel in enumerate(channels, start=2):
                values = list(self.trend_data[channel])
                if idx < len(values):
                    ws_trend.cell(row=idx + 2, column=col, value=values[idx])
        self._autosize_sheet(ws_trend)

        wb.save(path)
        self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("excel_exported", path=path)}")
        self.log(f"Komplett-Excel exportiert: {path}")


    def export_trend_png(self):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"trend_{timestamp}.png"
        pixmap = self.trend_plot.grab()
        if pixmap.save(str(path)):
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("trend_png_exported", path=path)}")
            self.log(f"Trend PNG exportiert: {path}")
        else:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("trend_png_failed")}")

    def export_pdf_report(self):
        rows = self._sensor_rows_for_export()
        if not rows:
            self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("pdf_no_data")}")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = EXPORT_DIR / f"mcs4_report_{timestamp}.pdf"

        alarm_items = list(self.active_alarms.values())
        alarm_html = "".join(f"<li>{alarm}</li>" for alarm in alarm_items) or "<li>Keine aktiven Alarms</li>"

        sensor_rows = ""
        for sensor in rows:
            sensor_rows += (
                "<tr>"
                f"<td>{sensor['id']}</td>"
                f"<td>{sensor['name']}</td>"
                f"<td>{sensor['value']:.2f}</td>"
                f"<td>{sensor['unit']}</td>"
                f"<td>{sensor['state']}</td>"
                f"<td>{sensor['alarm']}</td>"
                f"<td>{sensor['min']:.2f}</td>"
                f"<td>{sensor['max']:.2f}</td>"
                f"<td>{sensor['time']}</td>"
                "</tr>"
            )

        html = f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 10pt; }}
                h1 {{ color: #1f2937; }}
                h2 {{ color: #374151; margin-top: 18px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #999; padding: 4px; }}
                th {{ background: #e5e7eb; }}
                .meta {{ color: #555; }}
            </style>
        </head>
        <body>
            <h1>MCS-4 Diagnosticsbericht</h1>
            <p class="meta">Erstellt: {datetime.now():%Y-%m-%d %H:%M:%S}</p>
            <p class="meta">Software version: {APP_VERSION}</p>
            <p class="meta">Modus: {self.mode.currentText()} | COM-Port: {self.port_box.currentData()}</p>
            <h2>Zusammenfassung</h2>
            <ul>
                <li>Dekodierte Telegrams: {self.packet_count}</li>
                <li>Bytes empfangen/erzeugt: {self.byte_count}</li>
                <li>Sensors in dashboard: {len(self.sensor_values)}</li>
                <li>Aktive Alarms: {len(self.active_alarms)}</li>
            </ul>
            <h2>Sensor Overview</h2>
            <table>
                <tr>
                    <th>ID</th><th>Sensor</th><th>Value</th><th>Unit</th><th>Status</th>
                    <th>Alarm</th><th>Min</th><th>Max</th><th>Zeit</th>
                </tr>
                {sensor_rows}
            </table>
            <h2>Aktive Alarms</h2>
            <ul>{alarm_html}</ul>
            <h2>Hinweis</h2>
            <p>The MCS-4 decoder is currently an extendable base decoder. Full word-type and scaling logic is being completed step by step using the MCS-4 documentation.</p>
        </body>
        </html>
        """

        document = QTextDocument()
        document.setHtml(html)
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(str(path))
        document.print_(printer)

        self.export_log.appendPlainText(f"{datetime.now():%H:%M:%S} {self.tr("pdf_exported", path=path)}")
        self.log(f"PDF-Bericht exportiert: {path}")

    def load_player_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "MCS-Logdatei laden",
            str(REC_DIR),
            "MCS Log (*.mcslog);;Alle Dateien (*.*)",
        )
        if not path:
            return

        self.player_file_path = Path(path)
        self.player_packets.clear()
        self.player_index = 0

        with self.player_file_path.open("r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 2:
                    hex_part = parts[-1]
                    try:
                        self.player_packets.append(bytes.fromhex(hex_part))
                    except ValueError:
                        pass

        self.mode.setCurrentText("Player")
        self.status.setText(self.tr("status_player_file_loaded", count=len(self.player_packets)))
        self.log(f"Player-Datei geladen: {self.player_file_path}")

    def start_player(self):
        if not self.player_packets:
            self.status.setText(self.tr("status_no_player_file"))
            self.log("Keine Player-Datei geladen")
            return

        self.player_index = 0
        self.status.setText(self.tr("status_player_running"))
        self.log("Player gestartet")
        self.player_timer.start(100)

    def player_tick(self):
        if self.player_index >= len(self.player_packets):
            self.player_timer.stop()
            self.status.setText(self.tr("status_player_finished"))
            self.log("Player fertig")
            return

        packet = self.player_packets[self.player_index]
        self.player_index += 1
        self.byte_count += len(packet)
        self.telegram_log.appendPlainText(f"PLAY {hex_string(packet)}")
        self.process_bytes(packet)


    def register_protocol_packet(self, packet: bytes, word_type: int, number: int, page, line, key, validation_errors: list[str]):
        """Collect live protocol statistics for reverse engineering.

        This does not affect the dashboard. It only counts what the bus sends so
        that unknown measuring points and unusual word types can be analyzed.
        """
        self.protocol_total_count += 1
        self.protocol_wordtype_counts[word_type] += 1
        self.protocol_mp_counts[number] += 1
        self.protocol_packet_times.append(datetime.now())
        if validation_errors:
            self.protocol_invalid_count += 1

    def register_protocol_data_key(self, measuring_point: int, page: int, line: int, unit: str, raw_value: int, value: float):
        key = sensor_key(measuring_point, page, line)
        self.protocol_key_counts[key] += 1

    def _telegram_rate_per_second(self) -> float:
        if len(self.protocol_packet_times) < 2:
            return 0.0
        now = datetime.now()
        recent = [t for t in self.protocol_packet_times if (now - t).total_seconds() <= 5]
        if len(recent) < 2:
            return 0.0
        span = max((recent[-1] - recent[0]).total_seconds(), 0.001)
        return len(recent) / span

    def update_protocol_statistics_tab(self):
        lines = []
        lines.append("MCS-4 Live Protocol Analyzer (PDF-konformer Header: WT=D3..D0, Flags=D7..D4)")
        lines.append("============================")
        lines.append(f"Zeit: {datetime.now():%Y-%m-%d %H:%M:%S}")
        lines.append(f"Gesamttelegramme: {self.protocol_total_count}")
        lines.append(f"Telegrams/s (letzte ca. 5 s): {self._telegram_rate_per_second():.1f}")
        lines.append(f"Ungültige Telegrams: {self.protocol_invalid_count}")
        lines.append("")
        lines.append("WordType-Verteilung:")
        if not self.protocol_wordtype_counts:
            lines.append("  Noch keine Telegrams erfasst")
        else:
            for wt, count in sorted(self.protocol_wordtype_counts.items()):
                label = self.word_types.get(wt, "unbekannt")
                lines.append(f"  WT {wt}: {count:8d}  {label}")
        lines.append("")
        lines.append("Most frequent measuring points / numbers:")
        if not self.protocol_mp_counts:
            lines.append("  Noch keine Measuring Pointe erfasst")
        else:
            for mp, count in sorted(self.protocol_mp_counts.items(), key=lambda x: (-x[1], x[0]))[:30]:
                lines.append(f"  MP/NO {mp:3d}: {count:8d}")
        lines.append("")
        lines.append("Most frequent Data Value keys (MP:Page:Line):")
        if not self.protocol_key_counts:
            lines.append("  Noch keine Data-Value-Keys erfasst")
        else:
            for key, count in sorted(self.protocol_key_counts.items(), key=lambda x: (-x[1], sort_sensor_id(x[0])))[:30]:
                name = self.sensor_names.get(key, "unbekannt")
                lines.append(f"  {key:10s}: {count:8d}  {name}")
        lines.append("")
        lines.append("Hinweis:")
        lines.append("  Diese Statistik hilft, echte zyklische Messwerte, seltene Meldungen")
        lines.append("  und unbekannte Telegrams auf einer realen MCS-4-Anlage zu erkennen.")
        self.protocol_stats_log.setPlainText("\n".join(lines))

    def update_diagnostics(self):
        self.update_protocol_statistics_tab()
        text = (
            f"Zeit: {datetime.now():%Y-%m-%d %H:%M:%S}\n"
            f"Modus: {self.mode.currentText()}\n"
            f"Telegrams dekodiert: {self.packet_count}\n"
            f"Bytes empfangen/erzeugt: {self.byte_count}\n"
            f"Aktueller COM-Port: {self.port_box.currentData()}\n"
            f"RS422 aktiv: {'Yes' if self.serial_port else 'No'}\n"
            f"Sensors in dashboard: {len(self.sensor_values)}\n"
            f"Aktive Alarms: {len(self.active_alarms)}\n"
            f"Recording active: {'Yes' if self.recording else 'No'}\n"
            f"Aufnahme-Telegrams: {self.record_packet_count}\n"
            f"Player-Datei: {self.player_file_path.name if self.player_file_path else '-'}\n"
            f"Configuration file: {CONFIG_FILE}\n"
            f"Export folder: {EXPORT_DIR}\n"
            f"Decoder: MCS-4 Basisdecoder, WordType/Uniten erweiterbar\n"
        )
        self.diagnostics.setPlainText(text)

    def log(self, text):
        self.telegram_log.appendPlainText(f"LOG {datetime.now():%H:%M:%S} {text}")

    def closeEvent(self, event):
        self.stop_recording()
        self.stop()
        event.accept()


def main():
    app = QApplication(sys.argv)

    license_status = LicenseManager().validate()
    if not license_status.valid:
        QMessageBox.critical(
            None,
            "License required",
            "MCS-4 Professional Monitor cannot start.\n\n"
            f"Reason: {license_status.reason}\n\n"
            f"This PC Machine ID:\n{machine_id()}\n\n"
            "Please request a valid license.mcs file for this PC."
        )
        sys.exit(2)

    if license_status.days_left <= 5:
        QMessageBox.warning(
            None,
            "License expires soon",
            f"The license for {license_status.customer} expires in {license_status.days_left} day(s)."
        )

    win = MainWindow(license_status)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
